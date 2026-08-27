"""税务导出 Excel 摄取测试（队列 #295，design：fi2-tax-export-ingest）——全程假连接器，不触网。"""
from __future__ import annotations

import csv
import json

import openpyxl
import pytest

import fi2.tax_export_ingest as tei
from fi2.tax_export_ingest import (
    discover_new_files,
    ensure_segments,
    ingest_directory,
    is_processed,
    load_ledger,
    mark_processed,
    next_seq,
    parse_export_workbook,
    resolve_ap_no,
    resolve_item_code,
    save_ledger,
    write_invoice_csv,
)

_HEADER = [
    "序号", "发票代码", "发票号码", "数电发票号码", "销方识别号", "销方名称",
    "购方识别号", "购买方名称", "开票日期", "税收分类编码", "特定业务类型",
    "货物或应税劳务名称", "规格型号", "单位", "数量", "单价", "金额", "税率",
    "税额", "价税合计", "发票来源", "发票票种", "发票状态", "是否正数发票",
    "发票风险等级", "开票人", "备注",
]


def _make_export_xlsx(path, rows, *, header=_HEADER, sheet_name="信息汇总表"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(header)
    for r in rows:
        ws.append(r)
    wb.save(path)


def _row(digital_no="26327000000742719331", name="*半导体*场效应管", spec="NVMJS2D5N06CLTWG",
         unit="个", qty=100, unit_price=25.0, amount=None, tax_rate="13%", tax_amount=None,
         date="2026-07-10 11:33:22"):
    amount = amount if amount is not None else round(qty * unit_price, 2)
    tax_amount = tax_amount if tax_amount is not None else round(amount * 0.13, 2)
    return ["1", "", "", digital_no, "SELLER-TAX-ID", "卖方公司", "BUYER-TAX-ID", "买方公司",
            date, "1060105010000000000", "", name, spec, unit, qty, unit_price, amount,
            tax_rate, tax_amount, round(amount + tax_amount, 2), "电子发票服务平台",
            "数电发票（增值税专用发票）", "正常", "是", "正常", "张三", "ZPCG20260101001"]


# ── Excel 解析（3.1）──────────────────────────────────────────────────────

def test_parse_export_workbook_reads_rows(tmp_path):
    p = tmp_path / "sample.xlsx"
    _make_export_xlsx(p, [_row(), _row(digital_no="26327000000742719332")])
    rows = parse_export_workbook(p)
    assert len(rows) == 2
    assert rows[0]["数电发票号码"] == "26327000000742719331"
    assert rows[1]["数电发票号码"] == "26327000000742719332"


def test_parse_export_workbook_accepts_suffixed_sheet_name(tmp_path):
    """队列 #82 真实数据核验：真实批量导出文件的 sheet 名是「信息汇总表1」（带后缀），
    与 round-1 验证样本的精确「信息汇总表」不同，须前缀匹配兼容、不得判为 parse_error。"""
    p = tmp_path / "sample.xlsx"
    _make_export_xlsx(p, [_row()], sheet_name="信息汇总表1")
    rows = parse_export_workbook(p)
    assert len(rows) == 1
    assert rows[0]["数电发票号码"] == "26327000000742719331"


def test_parse_export_workbook_missing_sheet_raises(tmp_path):
    p = tmp_path / "bad.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "别的表"
    wb.save(p)
    with pytest.raises(ValueError, match="信息汇总表"):
        parse_export_workbook(p)


def test_parse_export_workbook_missing_field_raises(tmp_path):
    p = tmp_path / "bad_field.xlsx"
    header = [h for h in _HEADER if h != "数电发票号码"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "信息汇总表"
    ws.append(header)
    wb.save(p)
    with pytest.raises(ValueError, match="数电发票号码"):
        parse_export_workbook(p)


def test_parse_export_workbook_skips_blank_rows(tmp_path):
    p = tmp_path / "sample.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "信息汇总表"
    ws.append(_HEADER)
    ws.append(_row())
    ws.append([None] * len(_HEADER))
    wb.save(p)
    rows = parse_export_workbook(p)
    assert len(rows) == 1


# ── 已处理清单（3.2）──────────────────────────────────────────────────────

def test_ledger_roundtrip_and_idempotency(tmp_path):
    ledger_path = tmp_path / "ledger.json"
    export_dir = tmp_path / "export"
    export_dir.mkdir()
    f1 = export_dir / "a.xlsx"
    _make_export_xlsx(f1, [_row()])

    ledger = load_ledger(ledger_path)
    assert ledger == {}

    new_files = discover_new_files(export_dir, ledger)
    assert len(new_files) == 1
    path, file_hash = new_files[0]
    mark_processed(ledger, file_hash, path.name, row_count=1, processed_at="2026-08-07T00:00:00Z")
    save_ledger(ledger_path, ledger)

    # 重跑：同内容文件应被跳过
    ledger2 = load_ledger(ledger_path)
    assert is_processed(ledger2, file_hash)
    assert discover_new_files(export_dir, ledger2) == []


def test_same_filename_different_content_is_new(tmp_path):
    ledger_path = tmp_path / "ledger.json"
    export_dir = tmp_path / "export"
    export_dir.mkdir()
    f1 = export_dir / "a.xlsx"
    _make_export_xlsx(f1, [_row()])
    ledger = load_ledger(ledger_path)
    _, h1 = discover_new_files(export_dir, ledger)[0]
    mark_processed(ledger, h1, "a.xlsx", row_count=1, processed_at="t1")
    save_ledger(ledger_path, ledger)

    # 同名文件内容变化（追加一行）→ 哈希变化 → 应被视为新文件
    _make_export_xlsx(f1, [_row(), _row(digital_no="26327000000742719999")])
    ledger2 = load_ledger(ledger_path)
    new_files = discover_new_files(export_dir, ledger2)
    assert len(new_files) == 1
    assert new_files[0][1] != h1


# ── ap_no 反查（3.3）──────────────────────────────────────────────────────

class _FakeApByInvoiceConnector:
    """模拟服务端 CONTAINS 语义：返回集合可能包含非真实候选（InvoiceNo 恰好含查询串
    但并非真正匹配），resolve_ap_no 必须靠客户端 suffix 校验滤掉噪声。"""

    def __init__(self, rows_by_query: dict[str, list[dict]]):
        self._rows_by_query = rows_by_query

    def get_ap_lines_by_invoice_no(self, invoice_no):
        return self._rows_by_query.get(invoice_no, [])


def test_resolve_ap_no_unique_hit_full_string_stored():
    digital_no = "26322000005633189671"
    conn = _FakeApByInvoiceConnector({
        digital_no[-8:]: [
            {"DocNo": "AP-2026070071", "InvoiceNo": digital_no},
            {"DocNo": "AP-2026070071", "InvoiceNo": digital_no},
        ],
    })
    ap_no, reason, detail = resolve_ap_no(conn, digital_no)
    assert ap_no == "AP-2026070071"
    assert reason == ""


def test_resolve_ap_no_unique_hit_suffix_only_stored():
    digital_no = "26327000000742719331"
    conn = _FakeApByInvoiceConnector({
        digital_no[-8:]: [{"DocNo": "AP-2026070036", "InvoiceNo": digital_no[-8:]}],
    })
    ap_no, reason, detail = resolve_ap_no(conn, digital_no)
    assert ap_no == "AP-2026070036"
    assert reason == ""


def test_resolve_ap_no_filters_out_contains_noise():
    """服务端 CONTAINS 噪声：候选行的 InvoiceNo 恰好含查询子串但不是右侧子串关系，
    客户端 suffix 校验必须把它滤掉，不得误采信。"""
    digital_no = "26322000005633189671"
    suffix = digital_no[-8:]
    conn = _FakeApByInvoiceConnector({
        suffix: [
            {"DocNo": "AP-REAL", "InvoiceNo": digital_no},              # 真正匹配
            {"DocNo": "AP-NOISE", "InvoiceNo": "9" + suffix + "9"},     # 噪声：非右侧子串
        ],
    })
    ap_no, reason, detail = resolve_ap_no(conn, digital_no)
    assert ap_no == "AP-REAL"
    assert reason == ""


def test_resolve_ap_no_zero_match():
    conn = _FakeApByInvoiceConnector({})
    ap_no, reason, detail = resolve_ap_no(conn, "26322000005633189671")
    assert ap_no is None
    assert reason == "ap_no_zero_match"


def test_resolve_ap_no_ambiguous_multiple_ap_docs():
    digital_no = "26322000005633189671"
    suffix = digital_no[-8:]
    conn = _FakeApByInvoiceConnector({
        suffix: [
            {"DocNo": "AP-A", "InvoiceNo": digital_no},
            {"DocNo": "AP-B", "InvoiceNo": suffix},
        ],
    })
    ap_no, reason, detail = resolve_ap_no(conn, digital_no)
    assert ap_no is None
    assert reason == "ap_no_ambiguous"
    assert "AP-A" in detail and "AP-B" in detail


# ── item_code 反查（3.4）──────────────────────────────────────────────────

def test_resolve_item_code_unique_match():
    ap_lines = [
        {"ItemCode": "R02E.0217", "APQtyTU": 183.0, "TaxPrice": 28.6},
        {"ItemCode": "R02E.0217", "APQtyTU": 157.0, "TaxPrice": 28.6},
        {"ItemCode": "R02E.0208", "APQtyTU": 250.0, "TaxPrice": 17.0},
    ]
    # 未税单价 25.3097345132743 * 1.13 ≈ 28.6
    item_code, reason, detail = resolve_item_code(ap_lines, qty=183.0, untaxed_unit_price=25.3097345132743, tax_rate=0.13)
    assert item_code == "R02E.0217"
    assert reason == ""


def test_resolve_item_code_zero_match_when_qty_not_in_ap_lines():
    """真实场景：发票把 AP 的多笔批次合并成一行（33+67=100），单笔 qty=100 在 AP
    明细中找不到匹配——必须如实标记未解析，不得猜测归到最接近的一行。"""
    ap_lines = [
        {"ItemCode": "R02E.0217", "APQtyTU": 33.0, "TaxPrice": 28.6},
        {"ItemCode": "R02E.0217", "APQtyTU": 67.0, "TaxPrice": 28.6},
    ]
    item_code, reason, detail = resolve_item_code(ap_lines, qty=100.0, untaxed_unit_price=25.3097345132743, tax_rate=0.13)
    assert item_code is None
    assert reason == "item_code_zero_match"


def test_resolve_item_code_ambiguous_same_qty_price_different_items():
    ap_lines = [
        {"ItemCode": "X001", "APQtyTU": 10.0, "TaxPrice": 5.0},
        {"ItemCode": "X002", "APQtyTU": 10.0, "TaxPrice": 5.0},
    ]
    item_code, reason, detail = resolve_item_code(ap_lines, qty=10.0, untaxed_unit_price=5.0 / 1.13, tax_rate=0.13)
    assert item_code is None
    assert reason == "item_code_ambiguous"
    assert "X001" in detail and "X002" in detail


# ── 端到端编排（3.5）──────────────────────────────────────────────────────

class _FakeFullConnector:
    def __init__(self, invoice_rows, ap_lines_by_ap_no):
        self._invoice_rows = invoice_rows
        self._ap_lines_by_ap_no = ap_lines_by_ap_no

    def get_ap_lines_by_invoice_no(self, invoice_no):
        return self._invoice_rows.get(invoice_no, [])

    def get_ap_lines(self, ap_no):
        return self._ap_lines_by_ap_no.get(ap_no, [])


def test_ingest_directory_end_to_end_produces_invoice_csv(tmp_path):
    export_dir = tmp_path / "export"
    export_dir.mkdir()
    digital_no = "26327000000742719331"
    _make_export_xlsx(export_dir / "one.xlsx", [
        _row(digital_no=digital_no, qty=10, unit_price=100.0, tax_rate="13%"),
    ])

    conn = _FakeFullConnector(
        invoice_rows={digital_no[-8:]: [{"DocNo": "AP-1", "InvoiceNo": digital_no}]},
        ap_lines_by_ap_no={"AP-1": [{"ItemCode": "X001", "APQtyTU": 10.0, "TaxPrice": 113.0}]},
    )

    ledger_path = tmp_path / "ledger.json"
    result = ingest_directory(export_dir, ledger_path, conn, now="2026-08-07T00:00:00Z")

    assert result.files_processed == ["one.xlsx"]
    assert result.files_skipped == []
    assert len(result.resolved_rows) == 1
    row = result.resolved_rows[0]
    assert row["ap_no"] == "AP-1"
    assert row["item_code"] == "X001"
    assert row["inv_no"] == digital_no
    assert row["inv_qty"] == 10
    assert result.diagnostics == []

    out_csv = tmp_path / "out" / "invoice.csv"
    write_invoice_csv(result.resolved_rows, out_csv)
    with open(out_csv, encoding="utf-8-sig") as f:
        written = list(csv.DictReader(f))
    assert len(written) == 1
    assert written[0]["ap_no"] == "AP-1"
    assert written[0]["item_code"] == "X001"

    # 幂等：同一目录再摄取一次，文件已在清单中，不重复产出
    result2 = ingest_directory(export_dir, ledger_path, conn, now="2026-08-07T01:00:00Z")
    assert result2.files_processed == []
    assert result2.files_skipped == ["one.xlsx"]
    assert result2.resolved_rows == []


def test_ingest_directory_unresolved_rows_are_diagnosed_not_dropped_silently(tmp_path):
    export_dir = tmp_path / "export"
    export_dir.mkdir()
    digital_no = "26327000000742719331"
    _make_export_xlsx(export_dir / "one.xlsx", [_row(digital_no=digital_no)])

    conn = _FakeFullConnector(invoice_rows={}, ap_lines_by_ap_no={})   # 零命中
    ledger_path = tmp_path / "ledger.json"
    result = ingest_directory(export_dir, ledger_path, conn, now="2026-08-07T00:00:00Z")

    assert result.resolved_rows == []
    assert len(result.diagnostics) == 1
    assert result.diagnostics[0].reason == "ap_no_zero_match"
    assert result.diagnostics[0].digital_invoice_no == digital_no
    # 文件仍计入已处理（`discover_new_files` 此后跳过它）。
    # ⚠️ 2026-08-26 队列 #418 更正此处的原有注释：当初写的是「该发票已被诊断过，不会
    # 每次重跑都重新报告同一未解析行」——**那正是那 4 张假「无发票支撑」的成因**。
    # 该行如今会被登记进 ledger 的 `unresolved` 并在后续每次运行重试，见本文件
    # `test_unresolved_row_is_retried_once_the_ap_doc_finally_exists`。
    assert result.files_processed == ["one.xlsx"]
    assert result.pending_unresolved == 1


def test_ingest_directory_parse_error_file_produces_file_level_diagnostic(tmp_path):
    export_dir = tmp_path / "export"
    export_dir.mkdir()
    wb = openpyxl.Workbook()
    wb.active.title = "别的表"
    wb.save(export_dir / "bad.xlsx")

    conn = _FakeFullConnector(invoice_rows={}, ap_lines_by_ap_no={})
    ledger_path = tmp_path / "ledger.json"
    result = ingest_directory(export_dir, ledger_path, conn, now="2026-08-07T00:00:00Z")

    assert result.resolved_rows == []
    assert len(result.diagnostics) == 1
    assert result.diagnostics[0].reason == "parse_error"
    assert result.diagnostics[0].row_index is None


def test_write_invoice_csv_appends_to_existing_file(tmp_path):
    out_csv = tmp_path / "invoice.csv"
    write_invoice_csv([{
        "inv_no": "INV-1", "ap_no": "AP-1", "item_code": "X001", "unit": "个",
        "unit_price": 1.0, "inv_qty": 10, "untaxed_amount": 8.85, "tax_rate": 0.13,
        "tax_amount": 1.15, "inv_date": "2026-01-01",
    }], out_csv)
    write_invoice_csv([{
        "inv_no": "INV-2", "ap_no": "AP-2", "item_code": "X002", "unit": "个",
        "unit_price": 2.0, "inv_qty": 5, "untaxed_amount": 8.85, "tax_rate": 0.13,
        "tax_amount": 1.15, "inv_date": "2026-01-02",
    }], out_csv)
    with open(out_csv, encoding="utf-8-sig") as f:
        rows = list(csv.DictReader(f))
    assert [r["inv_no"] for r in rows] == ["INV-1", "INV-2"]


# ── 发票级幂等闸（队列 #371，2026-08-24）────────────────────────────────────
#
# 背景：唐燕萍 2026-08-21 举证面板发票数字整整翻倍。根因＝去重只有「文件内容 SHA256」
# 一层，同一张发票出现在两份**字节不同**的 xlsx 里就会被摄取两次，`write_invoice_csv`
# 追加写，面板按 (ap_no, item_code) 聚合求和后翻倍。下列用例锁死修复后的行为。


def _dup_conn(digital_no, ap_no="AP-2026070036", item_code="R01F.0034",
              qty=21000.0, taxed_price=3.9098):
    return _FakeFullConnector(
        invoice_rows={digital_no[-8:]: [{"DocNo": ap_no, "InvoiceNo": digital_no}]},
        ap_lines_by_ap_no={ap_no: [{"ItemCode": item_code, "APQtyTU": qty,
                                     "TaxPrice": taxed_price}]},
    )


def test_same_invoice_in_two_byte_different_files_is_ingested_only_once(tmp_path):
    """🔴 #371 主用例：两份**字节不同**、但含同一张发票的 xlsx —— 摄取后行数不翻倍。

    这正是真实事故的形态：她重导/另存了一次，两份文件内容哈希不同，旧闸放行两次。
    断言用的是她举证的那张真实发票的真实数字（21000／72660／9445.8）。
    """
    from fi2.tax_export_ingest import _hash_file, load_ingested_invoice_nos

    export_dir = tmp_path / "export"
    export_dir.mkdir()
    digital_no = "26327000000742719331"
    inv_row = _row(digital_no=digital_no, qty=21000, unit_price=3.46,
                   amount=72660.0, tax_rate="13%", tax_amount=9445.8)

    _make_export_xlsx(export_dir / "a.xlsx", [inv_row])
    _make_export_xlsx(export_dir / "b.xlsx", [inv_row], sheet_name="信息汇总表1")
    # 两份文件的哈希确实不同——否则本用例测的不是它想测的东西（旧闸本就会挡住同哈希）
    assert _hash_file(export_dir / "a.xlsx") != _hash_file(export_dir / "b.xlsx")

    conn = _dup_conn(digital_no)
    out_csv = tmp_path / "out" / "invoice.csv"
    result = ingest_directory(export_dir, tmp_path / "ledger.json", conn,
                              now="2026-08-24T00:00:00Z",
                              known_invoice_nos=load_ingested_invoice_nos(out_csv))
    write_invoice_csv(result.resolved_rows, out_csv)

    assert sorted(result.files_processed) == ["a.xlsx", "b.xlsx"]
    assert len(result.resolved_rows) == 1, "同一张发票被摄取了两次——#371 复发"
    assert result.duplicate_rows_skipped == 1
    assert result.duplicate_invoice_nos == [digital_no]

    with open(out_csv, encoding="utf-8-sig") as f:
        written = list(csv.DictReader(f))
    assert len(written) == 1
    assert float(written[0]["inv_qty"]) == 21000.0
    assert float(written[0]["untaxed_amount"]) == 72660.0
    assert float(written[0]["tax_amount"]) == 9445.8


def test_duplicate_invoice_is_skipped_across_separate_runs(tmp_path):
    """跨**批次**同样要挡住：第二天的新文件里含昨天已入库的发票。

    真实链路就是这样——闸的状态来自 `invoice.csv` 自身，不是进程内变量。
    """
    from fi2.tax_export_ingest import load_ingested_invoice_nos

    export_dir = tmp_path / "export"
    export_dir.mkdir()
    digital_no = "26327000000742719331"
    inv_row = _row(digital_no=digital_no, qty=21000, unit_price=3.46,
                   amount=72660.0, tax_rate="13%", tax_amount=9445.8)
    _make_export_xlsx(export_dir / "day1.xlsx", [inv_row])

    conn = _dup_conn(digital_no)
    ledger_path = tmp_path / "ledger.json"
    out_csv = tmp_path / "out" / "invoice.csv"

    r1 = ingest_directory(export_dir, ledger_path, conn, now="2026-08-24T00:00:00Z",
                          known_invoice_nos=load_ingested_invoice_nos(out_csv))
    write_invoice_csv(r1.resolved_rows, out_csv)
    assert len(r1.resolved_rows) == 1

    # 第二天：区间重叠的新导出文件，含同一张发票
    _make_export_xlsx(export_dir / "day2.xlsx", [inv_row], sheet_name="信息汇总表1")
    r2 = ingest_directory(export_dir, ledger_path, conn, now="2026-08-25T00:00:00Z",
                          known_invoice_nos=load_ingested_invoice_nos(out_csv))
    write_invoice_csv(r2.resolved_rows, out_csv)

    assert r2.files_processed == ["day2.xlsx"]
    assert r2.resolved_rows == []
    assert r2.duplicate_rows_skipped == 1
    with open(out_csv, encoding="utf-8-sig") as f:
        assert len(list(csv.DictReader(f))) == 1


def test_repeated_identical_lines_within_ONE_file_are_all_kept(tmp_path):
    """🔴 反例锁死：同一份文件里同一张发票的多行**合法重复**，一行都不许删。

    真实数据实证（`.51`，2026-08-24）：`26942000000588188581` 单张发票 60 行，其中一组
    (料品, 数量, 单价) 完全相同的签名重复 6 次，全部来自同一份源文件。**按行内容去重
    会把这些合法行删掉** —— 故闸只看「这张发票是不是别的文件已经贡献过」，不看行内容。
    """
    from fi2.tax_export_ingest import load_ingested_invoice_nos

    export_dir = tmp_path / "export"
    export_dir.mkdir()
    digital_no = "26942000000588188581"
    same = _row(digital_no=digital_no, qty=100, unit_price=25.0, tax_rate="13%")
    _make_export_xlsx(export_dir / "one.xlsx", [same, same, same])

    conn = _dup_conn(digital_no, ap_no="AP-9", item_code="Y001",
                     qty=100.0, taxed_price=28.25)
    result = ingest_directory(export_dir, tmp_path / "ledger.json", conn,
                              now="2026-08-24T00:00:00Z",
                              known_invoice_nos=load_ingested_invoice_nos(tmp_path / "none.csv"))

    assert len(result.resolved_rows) == 3, "同一文件内的合法重复行被误删——闸开得太宽"
    assert result.duplicate_rows_skipped == 0


def test_known_invoice_nos_none_still_dedups_within_one_run(tmp_path):
    """⚠️ 不传 `known_invoice_nos` **不等于关掉闸**：本次运行内跨文件的重复照样挡住。

    传 None 只表示「没有历史包袱」。锁死这一点是为了防止后来者误以为「不传就是旧行为」
    而在别处依赖一个并不存在的开关；真正会漏的是**跨批次**（见下一个用例的对照）。
    """
    export_dir = tmp_path / "export"
    export_dir.mkdir()
    digital_no = "26327000000742719331"
    inv_row = _row(digital_no=digital_no, qty=10, unit_price=100.0, tax_rate="13%")
    _make_export_xlsx(export_dir / "a.xlsx", [inv_row])
    _make_export_xlsx(export_dir / "b.xlsx", [inv_row], sheet_name="信息汇总表1")

    conn = _dup_conn(digital_no, ap_no="AP-1", item_code="X001",
                     qty=10.0, taxed_price=113.0)
    result = ingest_directory(export_dir, tmp_path / "ledger.json", conn,
                              now="2026-08-24T00:00:00Z")
    assert len(result.resolved_rows) == 1
    assert result.duplicate_rows_skipped == 1


def test_known_invoice_nos_none_misses_cross_batch_duplicate(tmp_path):
    """🔴 对照用例：不传 `known_invoice_nos` 时**跨批次**重复会漏过——#371 正是此形态。

    这个用例是刻意留下的「反面锚点」：它证明调用方必须传闸状态，光靠进程内累积不够。
    两个真实调用点（`scripts/ingest_tax_export.py`、`fi2/tax_export_scan.scan_once`）
    都已传；若哪天有人去掉，本用例与上方跨批次用例会一起说明代价。
    """
    export_dir = tmp_path / "export"
    export_dir.mkdir()
    digital_no = "26327000000742719331"
    inv_row = _row(digital_no=digital_no, qty=10, unit_price=100.0, tax_rate="13%")
    _make_export_xlsx(export_dir / "day1.xlsx", [inv_row])

    conn = _dup_conn(digital_no, ap_no="AP-1", item_code="X001",
                     qty=10.0, taxed_price=113.0)
    ledger_path = tmp_path / "ledger.json"
    r1 = ingest_directory(export_dir, ledger_path, conn, now="2026-08-24T00:00:00Z")
    assert len(r1.resolved_rows) == 1

    _make_export_xlsx(export_dir / "day2.xlsx", [inv_row], sheet_name="信息汇总表1")
    r2 = ingest_directory(export_dir, ledger_path, conn, now="2026-08-25T00:00:00Z")
    assert len(r2.resolved_rows) == 1           # 漏过：新批次不知道昨天已入库
    assert r2.duplicate_rows_skipped == 0


def test_load_ingested_invoice_nos_reads_existing_csv(tmp_path):
    from fi2.tax_export_ingest import load_ingested_invoice_nos

    out_csv = tmp_path / "invoice.csv"
    # 文件不存在＝首次摄取，返回空集合而非报错（「还没开始」不是异常）
    assert load_ingested_invoice_nos(out_csv) == set()

    write_invoice_csv([
        {"inv_no": "INV-1", "ap_no": "AP-1", "item_code": "X", "unit": "个",
         "unit_price": 1.0, "inv_qty": 1, "untaxed_amount": 1.0, "tax_rate": 0.13,
         "tax_amount": 0.13, "inv_date": "2026-01-01"},
        {"inv_no": "INV-2", "ap_no": "AP-1", "item_code": "Y", "unit": "个",
         "unit_price": 2.0, "inv_qty": 1, "untaxed_amount": 2.0, "tax_rate": 0.13,
         "tax_amount": 0.26, "inv_date": "2026-01-02"},
    ], out_csv)
    assert load_ingested_invoice_nos(out_csv) == {"INV-1", "INV-2"}


def test_duplicate_skips_do_not_pollute_diagnostics(tmp_path):
    """跨文件重复**不得**进 diagnostics —— 否则会被 `tax_export_scan` 的文件级失败
    判定与「未解析需人工核对」两处误读，真正的失败信号会被重复量淹没。
    """
    from fi2.tax_export_ingest import load_ingested_invoice_nos

    export_dir = tmp_path / "export"
    export_dir.mkdir()
    digital_no = "26327000000742719331"
    inv_row = _row(digital_no=digital_no, qty=10, unit_price=100.0, tax_rate="13%")
    _make_export_xlsx(export_dir / "a.xlsx", [inv_row])
    _make_export_xlsx(export_dir / "b.xlsx", [inv_row], sheet_name="信息汇总表1")

    conn = _dup_conn(digital_no, ap_no="AP-1", item_code="X001",
                     qty=10.0, taxed_price=113.0)
    result = ingest_directory(export_dir, tmp_path / "ledger.json", conn,
                              now="2026-08-24T00:00:00Z",
                              known_invoice_nos=load_ingested_invoice_nos(tmp_path / "none.csv"))
    assert result.duplicate_rows_skipped == 1
    assert result.diagnostics == []


# ── 未解析行必须可重试（队列 #418）────────────────────────────────────────
#
# 唐燕萍 2026-08-26 随机抽 10 张 AP 单，**4 张被面板报「无发票支撑」而发票实际存在**。
# 她的四组对照里有两组的发票落在《…20260401-20260430》导出文件里、AP 单却是 8 月的。
#
# 这里复现的是**机制**，不是那四个真实单号（真实复现须连 U9C，见队列 #418 LAN 留步）：
# 发票 4 月开出时摄取跑过一次 → 那时 AP 单还没立 → `ap_no_zero_match` → 该行只落一条
# 打印即弃的诊断 → 而文件 SHA 已进 ledger ⇒ **此后永远跳过、永不重试** ⇒ 8 月 AP 单
# 立了账，这张发票仍然不在 `invoice.csv` 里 ⇒ 面板报「无发票支撑」。
#
# 🔑 危险不在它报错，在它报的是一个看起来完全合理的结论——「发票还没到」在账上天天
# 真实发生，没人会怀疑。**错误不产生任何信号。**

class _LateApConnector(_FakeFullConnector):
    """AP 单「过一阵子才立账」的假连接器——`arrive()` 之后才反查得到。"""

    def __init__(self, digital_no, ap_no, ap_lines):
        super().__init__(invoice_rows={}, ap_lines_by_ap_no={})
        self._digital_no = digital_no
        self._ap_no = ap_no
        self._ap_lines = ap_lines

    def arrive(self):
        self._invoice_rows = {self._digital_no[-8:]:
                              [{"DocNo": self._ap_no, "InvoiceNo": self._digital_no}]}
        self._ap_lines_by_ap_no = {self._ap_no: self._ap_lines}


_APRIL_FILE = "全量发票查询导出结果（20260401-20260430）.xlsx"


def _late_ap_setup(tmp_path, digital_no="26322000003204358531", ap_no="AP-2026080137"):
    """4 月那份导出文件 + 8 月才立账的 AP 单（她第 2/4 组对照的形状）。"""
    export_dir = tmp_path / "export"
    export_dir.mkdir()
    _make_export_xlsx(export_dir / _APRIL_FILE,
                      [_row(digital_no=digital_no, qty=10, unit_price=100.0, tax_rate="13%")])
    conn = _LateApConnector(digital_no, ap_no,
                            [{"ItemCode": "X001", "APQtyTU": 10.0, "TaxPrice": 113.0}])
    return export_dir, conn, tmp_path / "ledger.json", digital_no, ap_no


def test_unresolved_row_is_retried_once_the_ap_doc_finally_exists(tmp_path):
    """#418 主诉复现 + 修复验证：AP 单晚于发票立账时，那张发票最终必须进得来。"""
    export_dir, conn, ledger_path, digital_no, ap_no = _late_ap_setup(tmp_path)

    # ① 4 月摄取：AP 单还没立 → 零命中
    r1 = ingest_directory(export_dir, ledger_path, conn, now="2026-04-30T00:00:00Z")
    assert r1.resolved_rows == []
    assert [d.reason for d in r1.diagnostics] == ["ap_no_zero_match"]
    assert r1.files_processed == [_APRIL_FILE]
    assert r1.pending_unresolved == 1          # 已登记，留待重试——不是打印即弃

    # ② 8 月：AP 单立了账。文件早已在 ledger 里，`discover_new_files` 依旧跳过它——
    #    修复前故事到此为止，这张发票永远进不来。
    conn.arrive()
    r2 = ingest_directory(export_dir, ledger_path, conn, now="2026-08-26T00:00:00Z",
                          known_invoice_nos=set())
    assert r2.files_processed == []            # 确实没有当成新文件
    assert r2.files_skipped == [_APRIL_FILE]
    assert r2.retried_rows_resolved == 1       # 而是被重试 pass 捞了回来
    assert r2.retried_invoice_nos == [digital_no]
    assert len(r2.resolved_rows) == 1
    assert r2.resolved_rows[0]["ap_no"] == ap_no
    assert r2.resolved_rows[0]["inv_no"] == digital_no
    assert r2.pending_unresolved == 0
    assert r2.unretryable_unresolved == 0

    # ③ 已解开的行必须从重试队列里消掉，不能每天重查一次
    entry = next(iter(load_ledger(ledger_path).values()))
    assert entry["unresolved"] == []
    r3 = ingest_directory(export_dir, ledger_path, conn, now="2026-08-27T00:00:00Z")
    assert r3.retried_rows_resolved == 0
    assert r3.resolved_rows == []


def test_retry_disabled_reproduces_the_418_defect(tmp_path):
    """关掉重试即退回旧行为——把「修复前是什么样」也钉在测试里，防止悄悄退化。"""
    export_dir, conn, ledger_path, _digital_no, _ap_no = _late_ap_setup(tmp_path)
    ingest_directory(export_dir, ledger_path, conn, now="2026-04-30T00:00:00Z",
                     retry_unresolved=False)
    conn.arrive()
    r2 = ingest_directory(export_dir, ledger_path, conn, now="2026-08-26T00:00:00Z",
                          retry_unresolved=False)
    assert r2.resolved_rows == []              # ← 这就是她看到的那 4 张假「无发票支撑」
    assert r2.retried_rows_resolved == 0


def test_unresolved_rows_are_persisted_in_the_ledger_not_just_printed(tmp_path):
    """根因之一：诊断此前只进 stdout。它必须落盘，否则重试无从谈起。"""
    export_dir, conn, ledger_path, digital_no, _ = _late_ap_setup(tmp_path)
    ingest_directory(export_dir, ledger_path, conn, now="2026-04-30T00:00:00Z")
    entry = next(iter(load_ledger(ledger_path).values()))
    assert len(entry["unresolved"]) == 1
    assert entry["unresolved"][0]["digital_invoice_no"] == digital_no
    assert entry["unresolved"][0]["reason"] == "ap_no_zero_match"
    assert entry["unresolved"][0]["row_index"] == 2


def test_blank_invoice_no_is_diagnosed_but_never_queued_for_retry(tmp_path):
    """发票号本身为空永远不会自愈——不该占着重试队列每天重查一次。"""
    export_dir = tmp_path / "export"
    export_dir.mkdir()
    _make_export_xlsx(export_dir / "one.xlsx", [_row(digital_no="")])
    conn = _FakeFullConnector(invoice_rows={}, ap_lines_by_ap_no={})
    ledger_path = tmp_path / "ledger.json"
    r = ingest_directory(export_dir, ledger_path, conn, now="2026-08-26T00:00:00Z")
    assert [d.reason for d in r.diagnostics] == ["digital_invoice_no_missing"]
    assert r.pending_unresolved == 0
    assert next(iter(load_ledger(ledger_path).values()))["unresolved"] == []


def test_deleted_source_file_makes_rows_unretryable_and_says_so(tmp_path):
    """🔴 源文件没了＝那些发票永远进不来。必须出声——沉默正是本缺陷当初的潜伏方式。"""
    export_dir, conn, ledger_path, _digital_no, _ = _late_ap_setup(tmp_path)
    ingest_directory(export_dir, ledger_path, conn, now="2026-04-30T00:00:00Z")

    (export_dir / _APRIL_FILE).unlink()
    conn.arrive()
    r2 = ingest_directory(export_dir, ledger_path, conn, now="2026-08-26T00:00:00Z")
    assert r2.retried_rows_resolved == 0
    assert r2.unretryable_unresolved == 1
    assert r2.unretryable_files == [_APRIL_FILE]


def test_changed_source_file_is_not_retried_by_filename(tmp_path):
    """同名不同内容的文件不得拿来顶替重试——row_index 会指到别的行上去（不猜）。"""
    export_dir, conn, ledger_path, digital_no, _ = _late_ap_setup(tmp_path)
    ingest_directory(export_dir, ledger_path, conn, now="2026-04-30T00:00:00Z")

    _make_export_xlsx(export_dir / _APRIL_FILE, [   # 同名、内容不同（前面多了一行）
        _row(digital_no="26322000009999999999", qty=1, unit_price=1.0),
        _row(digital_no=digital_no, qty=10, unit_price=100.0),
    ])
    conn.arrive()
    r2 = ingest_directory(export_dir, ledger_path, conn, now="2026-08-26T00:00:00Z")
    assert r2.unretryable_unresolved == 1
    assert r2.unretryable_files == [_APRIL_FILE]
    # 但它作为「同名不同内容」的新文件仍会被正常摄取（既有 #295 行为不变）
    assert r2.files_processed == [_APRIL_FILE]
    assert [r["inv_no"] for r in r2.resolved_rows] == [digital_no]


def test_retry_does_not_re_add_an_invoice_another_file_already_contributed(tmp_path):
    """重试不得绕过发票级幂等闸（队列 #371）——否则修好一个错、放回另一个错。"""
    export_dir, conn, ledger_path, digital_no, _ = _late_ap_setup(tmp_path)
    ingest_directory(export_dir, ledger_path, conn, now="2026-04-30T00:00:00Z")
    conn.arrive()
    # 该发票此后已由别的文件进过库（invoice.csv 里已有）
    r2 = ingest_directory(export_dir, ledger_path, conn, now="2026-08-26T00:00:00Z",
                          known_invoice_nos={digital_no})
    assert r2.retried_rows_resolved == 0
    assert r2.resolved_rows == []
    assert r2.duplicate_rows_skipped == 1
    assert next(iter(load_ledger(ledger_path).values()))["unresolved"] == []


def test_retry_contribution_is_recorded_as_a_separate_segment(tmp_path):
    """重试行追加在 CSV 尾部、与该文件最初那段并不相邻——ledger 必须如实记两段，
    否则 `rebuild_invoice_csv` 的归属重建（闸②/闸③）会整体错位。"""
    export_dir, conn, ledger_path, digital_no, _ = _late_ap_setup(tmp_path)
    other = "26322000001111111111"
    _make_export_xlsx(export_dir / _APRIL_FILE, [
        _row(digital_no=other, qty=7, unit_price=50.0),
        _row(digital_no=digital_no, qty=10, unit_price=100.0),
    ])
    conn._invoice_rows = {other[-8:]: [{"DocNo": "AP-OLD", "InvoiceNo": other}]}
    conn._ap_lines_by_ap_no = {"AP-OLD": [{"ItemCode": "Y001", "APQtyTU": 7.0, "TaxPrice": 56.5}]}
    r1 = ingest_directory(export_dir, ledger_path, conn, now="2026-04-30T00:00:00Z")
    assert len(r1.resolved_rows) == 1 and r1.pending_unresolved == 1

    conn.arrive()
    r2 = ingest_directory(export_dir, ledger_path, conn, now="2026-08-26T00:00:00Z")
    assert r2.retried_rows_resolved == 1

    entry = next(iter(load_ledger(ledger_path).values()))
    assert [s["row_count"] for s in entry["segments"]] == [1, 1]
    assert entry["row_count"] == 2          # 闸①口径：row_count 恒为各段之和
    assert entry["segments"][0]["seq"] < entry["segments"][1]["seq"]


def test_ensure_segments_migrates_a_legacy_ledger_in_place():
    """`.51` 上的现存 ledger 没有 `segments`——须按原 (processed_at, 文件名) 次序补齐。"""
    ledger = {
        "hB": {"file": "b.xlsx", "row_count": 5, "processed_at": "2026-08-02T00:00:00Z"},
        "hA": {"file": "a.xlsx", "row_count": 3, "processed_at": "2026-08-01T00:00:00Z"},
    }
    ensure_segments(ledger)
    assert ledger["hA"]["segments"] == [{"seq": 0, "row_count": 3}]   # 更早 → 更小 seq
    assert ledger["hB"]["segments"] == [{"seq": 1, "row_count": 5}]
    before = json.dumps(ledger, sort_keys=True)
    ensure_segments(ledger)                  # 幂等
    assert json.dumps(ledger, sort_keys=True) == before
    assert next_seq(ledger) == 2


# ── item_code 反查的候选口径（队列 #424）───────────────────────────────────
#
# 🔴 下面两组夹具是**唐燕萍 2026-08-26 四组举证里那两组失败案的逐字数据**（取证见
# `docs/queue_131_418_全BLOCK查证与全量对照-2026-08-27.md` §2.4），不是构造的样本。
# 它们的作用是把「现状口径为什么挡住它们、各候选口径分别能不能解开」钉死成可回归的
# 断言——**包括那些候选口径同样解不开的**（那正是最容易被乐观带过去的部分）。

#: ① `AP-2026080041` 密封胶 —— **计量单位不同**：发票按包装件（20 支），AP 按容量
#: （6200 ML ＝ 20 × 310ML）。未税 655.12／税额 85.16 两侧一分不差，数量差 310 倍。
_CASE_SEALANT_AP_ROWS = [
    {"ItemCode": "R02D.0001", "APQtyTU": 6200.0, "TaxPrice": 0.1194,
     "NonTaxAmtTC": 655.12, "TaxAmtTC": 85.16},
]
_CASE_SEALANT_INV = dict(qty=20.0, untaxed_unit_price=32.756, tax_rate=0.13,
                          untaxed_amount=655.12, tax_amount=85.16)

#: ② `AP-2026080137` 气泡袋 —— **发票按合计开票、AP 按行拆分**：
#: 发票 6000 ＝ 1000 ＋ 5000（**跨两个料号**）；发票 7000 ＝ 4000 ＋ 3000（同料号两行）。
_CASE_BUBBLE_AP_ROWS = [
    {"ItemCode": "J02E.0024", "APQtyTU": 1000.0, "TaxPrice": 0.20,
     "NonTaxAmtTC": 176.99, "TaxAmtTC": 23.01},
    {"ItemCode": "R02E.0024", "APQtyTU": 5000.0, "TaxPrice": 0.20,
     "NonTaxAmtTC": 884.96, "TaxAmtTC": 115.04},
    {"ItemCode": "R02E.0016", "APQtyTU": 4000.0, "TaxPrice": 0.34,
     "NonTaxAmtTC": 1203.54, "TaxAmtTC": 156.46},
    {"ItemCode": "R02E.0016", "APQtyTU": 3000.0, "TaxPrice": 0.34,
     "NonTaxAmtTC": 902.65, "TaxAmtTC": 117.35},
]
_CASE_BUBBLE_INV_6000 = dict(qty=6000.0, untaxed_unit_price=0.176991, tax_rate=0.13,
                              untaxed_amount=1061.95, tax_amount=138.05)
_CASE_BUBBLE_INV_7000 = dict(qty=7000.0, untaxed_unit_price=0.300885, tax_rate=0.13,
                              untaxed_amount=2106.19, tax_amount=273.81)


def test_default_item_match_strategy_is_unchanged_and_is_the_only_switch():
    """🔴 退化守卫：默认口径必须仍是现状 `qty_price`。

    本项目 `.51` 的部署方式是「整包同步」（队列 #418 ⑻ 实测坐实）——合入 master 即
    等于早晚上生产，「留步不部署」守不住。**唯一守得住的是这个默认值**，故它值得一条
    单独的断言，而不是靠人记得别改。
    """
    assert tei._ITEM_MATCH_STRATEGY == "qty_price"


def test_unknown_item_match_strategy_raises_rather_than_falls_back():
    """未知口径必须报错——静默回落到默认值正是「工具静默回退」那一族的坑。"""
    with pytest.raises(ValueError, match="未知的 item_code 匹配口径"):
        resolve_item_code(_CASE_SEALANT_AP_ROWS, strategy="whatever", **_CASE_SEALANT_INV)


def test_case_sealant_current_strategy_drops_the_row():
    """① 现状口径挡掉密封胶那一行——这就是唐燕萍那张发票「有票却报无票」的来路。"""
    code, reason, _ = resolve_item_code(_CASE_SEALANT_AP_ROWS, **_CASE_SEALANT_INV)
    assert code is None and reason == "item_code_zero_match"


def test_case_sealant_loosening_tolerance_does_not_help():
    """🔴 关键否定结论：**放宽容差解决不了 ①** —— 数量差的是 310 倍，不是浮点噪声。

    把 (数量, 单价) 两个容差各放宽到 10%（比现值大 4~5 个数量级）仍然零命中。
    ⇒ 「把容差调大一点」这条看起来最省事的路，在这一类上是无效的。
    """
    code, reason, _ = resolve_item_code(
        _CASE_SEALANT_AP_ROWS, qty_rel_tol=0.1, price_rel_tol=0.1, **_CASE_SEALANT_INV)
    assert code is None and reason == "item_code_zero_match"


@pytest.mark.parametrize("strategy", ["amount", "qty_price_then_amount",
                                       "qty_price_then_single_item"])
def test_case_sealant_solved_by_amount_or_single_item(strategy):
    """① 换成 (未税金额, 税额) 或「单料号回落」都能唯一解出——两侧金额一分不差。"""
    code, reason, _ = resolve_item_code(
        _CASE_SEALANT_AP_ROWS, strategy=strategy, **_CASE_SEALANT_INV)
    assert code == "R02D.0001" and reason == ""


def test_case_bubble_6000_not_solved_by_amount():
    """🔴 ② 的 6000 那一行：**⒜（金额口径）同样解不开**。

    队列 #424 原行写「两个真实案例里未税金额与税额都分毫不差」——那说的是**合计**：
    1000 ＋ 5000 两行加起来才等于发票那一行。逐行比时没有任何单独一行 AP 的金额等于
    1061.95 ⇒ 金额口径在这一类上和数量口径一样零命中。**不得把 ⒜ 当成通解。**
    """
    code, reason, _ = resolve_item_code(
        _CASE_BUBBLE_AP_ROWS, strategy="qty_price_then_amount", **_CASE_BUBBLE_INV_6000)
    assert code is None and reason == "item_code_zero_match"


def test_case_bubble_6000_subset_sum_is_ambiguous_because_it_spans_two_item_codes():
    """🔴 ② 的 6000 那一行：⒞ 凑得出来，但跨两个料号 ⇒ 如实报歧义，**不挑一个**。

    要真正解开这一类，必须同时接受「把一张发票行拆成多行写进 `invoice.csv`」——
    那是另一个口径决定（唐燕萍签认），不是本层能替她做的。
    """
    code, reason, detail = resolve_item_code(
        _CASE_BUBBLE_AP_ROWS, strategy="qty_price_then_subset_sum", **_CASE_BUBBLE_INV_6000)
    assert code is None and reason == "item_code_ambiguous"
    assert "J02E.0024" in detail and "R02E.0024" in detail


def test_case_bubble_7000_solved_by_subset_sum():
    """② 的 7000 那一行：4000 ＋ 3000 同属 `R02E.0016` ⇒ ⒞ 能唯一解出。"""
    code, reason, _ = resolve_item_code(
        _CASE_BUBBLE_AP_ROWS, strategy="qty_price_then_subset_sum", **_CASE_BUBBLE_INV_7000)
    assert code == "R02E.0016" and reason == ""


def test_case_bubble_single_item_fallback_does_not_apply_to_multi_code_doc():
    """⒝ 的「单料号回落」在多料号单据上必须不生效——否则就是随手挑一个。"""
    code, reason, _ = resolve_item_code(
        _CASE_BUBBLE_AP_ROWS, strategy="qty_price_then_single_item", **_CASE_BUBBLE_INV_6000)
    assert code is None and reason == "item_code_zero_match"


def test_item_match_fallback_never_fires_on_ambiguous_only_on_zero_match():
    """🔴 回落只在**零命中**时发生，绝不在**歧义**时发生。

    现状口径已经找到多个候选 ⇒ 这张发票行本身就分不清挂哪一行；换把尺子只会换一批
    候选、不会让它变清楚。那种「换到能出一个答案为止」正是静默猜测。
    """
    ap_rows = [
        {"ItemCode": "A1", "APQtyTU": 10.0, "TaxPrice": 1.13,
         "NonTaxAmtTC": 10.0, "TaxAmtTC": 1.3},
        {"ItemCode": "A2", "APQtyTU": 10.0, "TaxPrice": 1.13,
         "NonTaxAmtTC": 99.0, "TaxAmtTC": 9.9},
    ]
    inv = dict(qty=10.0, untaxed_unit_price=1.0, tax_rate=0.13,
               untaxed_amount=99.0, tax_amount=9.9)
    assert resolve_item_code(ap_rows, **inv)[1] == "item_code_ambiguous"
    # 加了金额回落也仍是歧义——不因为「金额只有 A2 对得上」就悄悄改判成 A2。
    assert resolve_item_code(ap_rows, strategy="qty_price_then_amount", **inv)[1] \
        == "item_code_ambiguous"


def test_amount_strategy_requires_both_amount_fields_and_never_guesses():
    """缺「金额」或「税额」时金额口径退化为零命中，**不由单价反推**。"""
    inv = dict(_CASE_SEALANT_INV)
    inv["tax_amount"] = None
    code, reason, _ = resolve_item_code(_CASE_SEALANT_AP_ROWS, strategy="amount", **inv)
    assert code is None and reason == "item_code_zero_match"


def test_subset_sum_solves_the_merged_batch_case_already_in_this_suite():
    """本文件早先那条「发票把 33 ＋ 67 合并成 100」的用例，在 ⒞ 下可唯一解开。

    ⚠️ 这不是说 ⒞ 就该采用：那条用例的注释写的是「必须如实标记未解析，不得猜测归到
    最接近的一行」——**⒞ 不是猜，它要求合计分毫不差**，但「合计相等即认定为同一笔」
    本身仍是一条账务口径，须唐燕萍签认。
    """
    ap_lines = [
        {"ItemCode": "R02E.0217", "APQtyTU": 33.0, "TaxPrice": 28.6},
        {"ItemCode": "R02E.0217", "APQtyTU": 67.0, "TaxPrice": 28.6},
    ]
    code, reason, _ = resolve_item_code(
        ap_lines, qty=100.0, untaxed_unit_price=25.3097345132743, tax_rate=0.13,
        strategy="qty_price_then_subset_sum")
    assert code == "R02E.0217" and reason == ""


# ── 丢行可见并可查（队列 #424）─────────────────────────────────────────────

def test_item_match_diagnosis_labels_the_dropped_row_without_amounts():
    """诊断串必须能把丢行分类，且**不含金额/数量原始值**（FI2 审计口径＝金额不落盘）。"""
    detail = tei._item_match_diagnosis(
        _CASE_SEALANT_AP_ROWS, ap_no="AP-2026080041", qty=20.0,
        untaxed_unit_price=32.756, tax_rate=0.13,
        untaxed_amount=655.12, tax_amount=85.16)
    assert "ap=AP-2026080041" in detail and "ap行数=1" in detail and "料号数=1" in detail
    assert "amount" in detail and "single_item" in detail
    for leaked in ("655.12", "85.16", "32.756", "6200"):
        assert leaked not in detail


def test_item_match_diagnosis_marks_multi_code_subset_sum_separately():
    """凑得出来但跨料号的，必须标成另一类，**不混进「可解」**。"""
    detail = tei._item_match_diagnosis(
        _CASE_BUBBLE_AP_ROWS, ap_no="AP-2026080137", **_CASE_BUBBLE_INV_6000)
    assert "subset_sum_跨2料号" in detail


def test_dropped_rows_carry_diagnosis_through_ingest(tmp_path):
    """端到端：一行挂不上料号的发票，其诊断必须带上可分类的身份并进重试队列。"""
    export_dir = tmp_path / "exports"
    export_dir.mkdir()
    _make_export_xlsx(export_dir / "a.xlsx", [
        _row(digital_no="26322000006465433531", qty=20, unit_price=32.756,
             amount=655.12, tax_amount=85.16),
    ])

    class _Conn:
        def get_ap_lines_by_invoice_no(self, suffix):
            return [{"InvoiceNo": "26322000006465433531", "DocNo": "AP-2026080041"}]

        def get_ap_lines(self, ap_no):
            return list(_CASE_SEALANT_AP_ROWS)

    result = ingest_directory(export_dir, tmp_path / "l.json", _Conn(),
                              now="2026-08-28T00:00:00Z")
    assert result.resolved_rows == []
    assert len(result.diagnostics) == 1
    d = result.diagnostics[0]
    assert d.reason == "item_code_zero_match"
    assert "ap=AP-2026080041" in d.detail and "换口径可解=" in d.detail


def test_summarize_and_write_diagnostics_jsonl(tmp_path):
    """未解析记录必须能汇总、能落盘 —— 补上 #418 根因链的第 ⑵ 环（诊断从不落盘）。"""
    result = tei.IngestResult(diagnostics=[
        tei.IngestDiagnostic(file="a.xlsx", reason="ap_no_zero_match",
                             digital_invoice_no="1", row_index=2),
        tei.IngestDiagnostic(file="a.xlsx", reason="ap_no_zero_match",
                             digital_invoice_no="1", row_index=3),
        tei.IngestDiagnostic(file="a.xlsx", reason="item_code_zero_match",
                             digital_invoice_no="2", row_index=4, detail="ap=AP-1"),
    ])
    assert tei.summarize_diagnostics(result) == [
        ("ap_no_zero_match", 2, 1), ("item_code_zero_match", 1, 1)]
    p = tmp_path / "diag.jsonl"
    assert tei.write_diagnostics_jsonl(result, p, now="2026-08-28T00:00:00Z") == 3
    # 追加写、不覆盖：第二次跑完是 6 行，历史查得回去。
    tei.write_diagnostics_jsonl(result, p, now="2026-08-28T01:00:00Z")
    lines = p.read_text(encoding="utf-8").strip().splitlines()
    assert len(lines) == 6
    rec = json.loads(lines[2])
    assert rec["reason"] == "item_code_zero_match" and rec["detail"] == "ap=AP-1"
