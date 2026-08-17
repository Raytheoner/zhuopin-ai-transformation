"""税务导出定时扫描 + 失败告警测试（队列 #82 第2层，2026-08-10）——全程假连接器/假
发送函数，不触网、不触发真实企微 webhook。复用 test_tax_export_ingest.py 的 Excel
构造辅助，验证 Layer 2 只加"文件级失败判定+告警"，不改 Layer 1 的摄取结果本身。"""
from __future__ import annotations

import csv

import openpyxl
import pytest

from fi2.tax_export_scan import (
    ScanFailedError,
    build_alert_message,
    file_level_failures,
    scan_once,
)

_HEADER = [
    "序号", "发票代码", "发票号码", "数电发票号码", "销方识别号", "销方名称",
    "购方识别号", "购买方名称", "开票日期", "税收分类编码", "特定业务类型",
    "货物或应税劳务名称", "规格型号", "单位", "数量", "单价", "金额", "税率",
    "税额", "价税合计", "发票来源", "发票票种", "发票状态", "是否正数发票",
    "发票风险等级", "开票人", "备注",
]


def _make_export_xlsx(path, rows, *, header=_HEADER, sheet_name="信息汇总表"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(header)
    for r in rows:
        ws.append(r)
    wb.save(path)


def _row(digital_no="26327000000742719331", name="*半导体*场效应管", spec="NVMJS2D5N06CLTWG",
         unit="个", qty=100, unit_price=25.0, amount=None, tax_rate="13%", tax_amount=None,
         date="2026-07-10 11:33:22"):
    amount = amount if amount is not None else round(qty * unit_price, 2)
    tax_amount = tax_amount if tax_amount is not None else round(amount * 0.13, 2)
    return ["1", "", "", digital_no, "SELLER-TAX-ID", "卖方公司", "BUYER-TAX-ID", "买方公司",
            date, "1060105010000000000", "", name, spec, unit, qty, unit_price, amount,
            tax_rate, tax_amount, round(amount + tax_amount, 2), "电子发票服务平台",
            "数电发票（增值税专用发票）", "正常", "是", "正常", "张三", "ZPCG20260101001"]


class _FakeFullConnector:
    def __init__(self, invoice_rows, ap_lines_by_ap_no):
        self._invoice_rows = invoice_rows
        self._ap_lines_by_ap_no = ap_lines_by_ap_no

    def get_ap_lines_by_invoice_no(self, invoice_no):
        return self._invoice_rows.get(invoice_no, [])

    def get_ap_lines(self, ap_no):
        return self._ap_lines_by_ap_no.get(ap_no, [])


class _RaisingConnector:
    """模拟 U9C 连接/网络故障——`get_ap_lines_by_invoice_no` 直接抛异常。"""
    def get_ap_lines_by_invoice_no(self, invoice_no):
        raise ConnectionError("模拟：U9C 端点不可达")

    def get_ap_lines(self, ap_no):
        raise ConnectionError("模拟：U9C 端点不可达")


class _RecordingSender:
    """假发送函数，记录调用参数，不触网。"""
    def __init__(self, *, fail=False):
        self.calls: list[tuple[str, str]] = []
        self._fail = fail

    def __call__(self, webhook_url: str, content: str) -> None:
        self.calls.append((webhook_url, content))
        if self._fail:
            raise RuntimeError("模拟：webhook 发送失败")


# ── file_level_failures 判据（区分文件级 vs 行级）───────────────────────────

def test_file_level_failures_filters_out_row_level_diagnostics(tmp_path):
    export_dir = tmp_path / "export"
    export_dir.mkdir()
    good_digital_no = "26327000000742719331"
    _make_export_xlsx(export_dir / "good.xlsx", [_row(digital_no=good_digital_no)])
    bad_wb_path = export_dir / "bad_sheet.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "别的表"
    wb.save(bad_wb_path)

    conn = _FakeFullConnector(invoice_rows={}, ap_lines_by_ap_no={})  # 零命中 → 行级诊断
    outcome = scan_once(export_dir, tmp_path / "out", tmp_path / "ledger.json", conn,
                         now="2026-08-10T00:00:00Z")

    # 两份文件都被处理：good.xlsx 产生行级诊断（ap_no_zero_match，非文件级失败），
    # bad_sheet.xlsx 产生文件级诊断（parse_error）——只有后者计入 file_level_failures。
    reasons = {d.reason for d in outcome.result.diagnostics}
    assert "ap_no_zero_match" in reasons
    assert "parse_error" in reasons
    assert len(outcome.file_level_failures) == 1
    assert outcome.file_level_failures[0].reason == "parse_error"
    assert outcome.file_level_failures[0].file == "bad_sheet.xlsx"


def test_file_level_failures_empty_when_only_row_level_noise(tmp_path):
    """真实数据里约 89% 属预期噪声（队列 #82 08-10 回填）——大量行级零命中不应
    被误判为需要告警的"扫描失败"。"""
    export_dir = tmp_path / "export"
    export_dir.mkdir()
    _make_export_xlsx(export_dir / "one.xlsx", [
        _row(digital_no="26327000000742719331"),
        _row(digital_no="26327000000742719332"),
    ])
    conn = _FakeFullConnector(invoice_rows={}, ap_lines_by_ap_no={})
    outcome = scan_once(export_dir, tmp_path / "out", tmp_path / "ledger.json", conn,
                         now="2026-08-10T00:00:00Z")
    assert len(outcome.result.diagnostics) == 2
    assert outcome.file_level_failures == []
    assert outcome.alert_sent is False


# ── 告警触发 ────────────────────────────────────────────────────────────

def test_alert_sent_on_file_level_failure_when_webhook_configured(tmp_path):
    export_dir = tmp_path / "export"
    export_dir.mkdir()
    wb = openpyxl.Workbook()
    wb.active.title = "信息汇总表1完全不对的名字"
    wb.save(export_dir / "bad.xlsx")

    sender = _RecordingSender()
    conn = _FakeFullConnector(invoice_rows={}, ap_lines_by_ap_no={})
    outcome = scan_once(export_dir, tmp_path / "out", tmp_path / "ledger.json", conn,
                         now="2026-08-10T00:00:00Z",
                         webhook_url="https://example.invalid/webhook", sender=sender)

    assert len(outcome.file_level_failures) == 1
    assert outcome.alert_sent is True
    assert outcome.alert_error == ""
    assert len(sender.calls) == 1
    webhook_url, content = sender.calls[0]
    assert webhook_url == "https://example.invalid/webhook"
    assert "bad.xlsx" in content
    assert "parse_error" in content


def test_no_alert_attempt_when_webhook_not_configured(tmp_path):
    export_dir = tmp_path / "export"
    export_dir.mkdir()
    wb = openpyxl.Workbook()
    wb.active.title = "不对的表名"
    wb.save(export_dir / "bad.xlsx")

    sender = _RecordingSender()
    conn = _FakeFullConnector(invoice_rows={}, ap_lines_by_ap_no={})
    outcome = scan_once(export_dir, tmp_path / "out", tmp_path / "ledger.json", conn,
                         now="2026-08-10T00:00:00Z", webhook_url=None, sender=sender)

    assert len(outcome.file_level_failures) == 1
    assert outcome.alert_sent is False
    assert outcome.alert_error == ""
    assert sender.calls == []


def test_alert_send_failure_does_not_crash_scan(tmp_path):
    """告警通道自身故障（如网络问题）不得掩盖/中断扫描主流程——扫描结果仍完整返回。"""
    export_dir = tmp_path / "export"
    export_dir.mkdir()
    wb = openpyxl.Workbook()
    wb.active.title = "不对的表名"
    wb.save(export_dir / "bad.xlsx")

    sender = _RecordingSender(fail=True)
    conn = _FakeFullConnector(invoice_rows={}, ap_lines_by_ap_no={})
    outcome = scan_once(export_dir, tmp_path / "out", tmp_path / "ledger.json", conn,
                         now="2026-08-10T00:00:00Z",
                         webhook_url="https://example.invalid/webhook", sender=sender)

    assert len(outcome.file_level_failures) == 1
    assert outcome.alert_sent is False
    assert "模拟：webhook 发送失败" in outcome.alert_error
    # 扫描本身的产出未受影响（parse_error 文件本身不计入 files_processed，
    # 留待文件修复后下次重跑——同 test_tax_export_ingest.py 既有行为约定）
    assert outcome.result.diagnostics[0].reason == "parse_error"


def test_no_failures_no_alert_attempt(tmp_path):
    export_dir = tmp_path / "export"
    export_dir.mkdir()
    digital_no = "26327000000742719331"
    _make_export_xlsx(export_dir / "one.xlsx", [_row(digital_no=digital_no, qty=10, unit_price=100.0)])
    conn = _FakeFullConnector(
        invoice_rows={digital_no[-8:]: [{"DocNo": "AP-1", "InvoiceNo": digital_no}]},
        ap_lines_by_ap_no={"AP-1": [{"ItemCode": "X001", "APQtyTU": 10.0, "TaxPrice": 113.0}]},
    )
    sender = _RecordingSender()
    outcome = scan_once(export_dir, tmp_path / "out", tmp_path / "ledger.json", conn,
                         now="2026-08-10T00:00:00Z",
                         webhook_url="https://example.invalid/webhook", sender=sender)

    assert outcome.file_level_failures == []
    assert outcome.alert_sent is False
    assert sender.calls == []
    out_csv = tmp_path / "out" / "invoice.csv"
    assert out_csv.is_file()
    with open(out_csv, encoding="utf-8-sig") as f:
        rows = list(csv.DictReader(f))
    assert len(rows) == 1
    assert rows[0]["ap_no"] == "AP-1"


# ── 扫描本身异常（连接/目录不可达）── ────────────────────────────────────

def test_scan_error_raises_scan_failed_error_and_still_alerts(tmp_path):
    export_dir = tmp_path / "export"
    export_dir.mkdir()
    _make_export_xlsx(export_dir / "one.xlsx", [_row()])

    sender = _RecordingSender()
    with pytest.raises(ScanFailedError) as exc_info:
        scan_once(export_dir, tmp_path / "out", tmp_path / "ledger.json", _RaisingConnector(),
                  now="2026-08-10T00:00:00Z",
                  webhook_url="https://example.invalid/webhook", sender=sender)

    outcome = exc_info.value.outcome
    assert outcome.result is None
    assert "U9C 端点不可达" in outcome.scan_error
    assert outcome.alert_sent is True
    assert len(sender.calls) == 1
    assert "U9C 端点不可达" in sender.calls[0][1]


def test_scan_error_without_webhook_still_raises(tmp_path):
    export_dir = tmp_path / "export"
    export_dir.mkdir()
    _make_export_xlsx(export_dir / "one.xlsx", [_row()])

    with pytest.raises(ScanFailedError) as exc_info:
        scan_once(export_dir, tmp_path / "out", tmp_path / "ledger.json", _RaisingConnector(),
                  now="2026-08-10T00:00:00Z", webhook_url=None)

    assert exc_info.value.outcome.alert_sent is False
    assert exc_info.value.outcome.alert_error == ""


# ── 告警正文组装 ────────────────────────────────────────────────────────

def test_build_alert_message_lists_failures_and_truncates(tmp_path):
    from fi2.tax_export_ingest import IngestDiagnostic
    failures = [
        IngestDiagnostic(file=f"f{i}.xlsx", reason="parse_error", detail=f"缺列{i}")
        for i in range(12)
    ]
    msg = build_alert_message(export_dir="D:\\airead", failures=failures)
    assert "文件级摄取失败 12 处" in msg
    assert "f0.xlsx" in msg
    assert "另有 2 处" in msg


def test_build_alert_message_scan_error_takes_priority(tmp_path):
    msg = build_alert_message(export_dir="D:\\airead", failures=[], scan_error="连接超时")
    assert "扫描本身未能完成" in msg
    assert "连接超时" in msg


# ── 源头断供检测（队列 #82，2026-08-17 补；见 tax_export_scan docstring ③）──
#
# 本组用例复现的是 2026-08-17 在 `.51` 上实测到的真实状态：`D:\airead` 自
# 2026-08-10 起 5 个工作日零新文件，扫描每天照跑、退出码恒 0、零诊断、零告警，
# 整整一周无人知晓。以下用例锁定"这种情况现在会说话"。

import datetime

from fi2.tax_export_scan import (
    SILENCE_WORKDAYS_DEFAULT,
    build_silence_message,
    detect_source_silence,
    last_ingest_at,
    workdays_between,
)


def _ledger(*stamps):
    return {f"hash{i}": {"file": f"f{i}.xlsx", "row_count": 1, "processed_at": s}
            for i, s in enumerate(stamps)}


def test_last_ingest_at_takes_max_not_last_inserted():
    # dict 插入序刻意与时间序不一致——取的必须是最大值，不是最后一条
    ledger = _ledger("2026-08-10T13:11:46+00:00", "2026-08-07T02:59:37+00:00")
    assert last_ingest_at(ledger) == "2026-08-10T13:11:46+00:00"


def test_last_ingest_at_empty_ledger():
    assert last_ingest_at({}) == ""


def test_last_ingest_at_ignores_unparsable_stamps():
    ledger = _ledger("2026-08-07T02:59:37+00:00", "not-a-timestamp")
    assert last_ingest_at(ledger) == "2026-08-07T02:59:37+00:00"


def test_workdays_between_skips_weekend():
    # 2026-08-14 周五 → 2026-08-17 周一：中间只有周一这一个工作日
    fri = datetime.datetime(2026, 8, 14, tzinfo=datetime.timezone.utc)
    mon = datetime.datetime(2026, 8, 17, tzinfo=datetime.timezone.utc)
    assert workdays_between(fri, mon) == 1


def test_workdays_between_real_incident_span():
    """2026-08-10（周一，最后一次真实摄取）→ 2026-08-17（周一，本次取证日）＝ 5 个工作日。"""
    start = datetime.datetime(2026, 8, 10, 13, 11, 46, tzinfo=datetime.timezone.utc)
    end = datetime.datetime(2026, 8, 17, 2, 30, tzinfo=datetime.timezone.utc)
    assert workdays_between(start, end) == 5


def test_workdays_between_same_day_and_backwards_are_zero():
    d = datetime.datetime(2026, 8, 17, tzinfo=datetime.timezone.utc)
    earlier = datetime.datetime(2026, 8, 10, tzinfo=datetime.timezone.utc)
    assert workdays_between(d, d) == 0
    assert workdays_between(d, earlier) == 0


def test_detect_source_silence_real_incident():
    """真实事故复现：08-10 最后摄取，08-17 判定 → 5 个工作日 ≥ 阈值 3，命中。"""
    silence = detect_source_silence(
        _ledger("2026-08-10T13:11:46+00:00"), "2026-08-17T02:30:00+00:00",
        SILENCE_WORKDAYS_DEFAULT)
    assert silence is not None
    assert silence.workdays_silent == 5
    assert silence.threshold == 3
    assert silence.last_ingest_at == "2026-08-10T13:11:46+00:00"


def test_detect_source_silence_below_threshold():
    # 08-14 周五 → 08-17 周一 ＝ 1 个工作日，未超阈值
    assert detect_source_silence(
        _ledger("2026-08-14T09:00:00+00:00"), "2026-08-17T02:30:00+00:00", 3) is None


def test_detect_source_silence_not_on_weekend():
    """周末跑到的扫描一律不告警——她的投放口径是工作日，周末没新文件是正常的。"""
    assert detect_source_silence(
        _ledger("2026-08-10T13:11:46+00:00"), "2026-08-15T02:30:00+00:00", 3) is None


def test_detect_source_silence_empty_ledger_is_not_silence():
    """从未摄取过 ≠ 断供——"还没开始"与"停了"性质不同，不混用同一条告警。"""
    assert detect_source_silence({}, "2026-08-17T02:30:00+00:00", 3) is None


def test_scan_alerts_on_source_silence_when_no_new_files(tmp_path):
    """端到端：目录里的文件全部已处理过（零新文件）+ 超阈值 → 发断供告警。"""
    export_dir = tmp_path / "export"
    export_dir.mkdir()
    _make_export_xlsx(export_dir / "one.xlsx", [_row()])
    ledger_path = tmp_path / "ledger.json"

    conn = _FakeFullConnector({"26327000000742719331": [{"DocNo": "AP-1", "InvoiceNo": "42719331"}]},
                              {"AP-1": []})
    # 第一次：真实摄取，写入 ledger（processed_at = 08-10）
    first = scan_once(export_dir, tmp_path / "out", ledger_path, conn,
                      now="2026-08-10T13:11:46+00:00", webhook_url="https://example.invalid/w",
                      sender=_RecordingSender())
    assert first.result.files_processed == ["one.xlsx"]
    assert first.source_silence is None            # 刚吃到新文件，不可能判断供

    # 第二次：同一目录零新文件，08-17 再扫 → 5 个工作日静默，命中
    sender = _RecordingSender()
    outcome = scan_once(export_dir, tmp_path / "out", ledger_path, conn,
                        now="2026-08-17T02:30:00+00:00",
                        webhook_url="https://example.invalid/w", sender=sender)

    assert outcome.result.files_processed == []
    assert outcome.source_silence is not None
    assert outcome.source_silence.workdays_silent == 5
    assert outcome.alert_sent is True
    assert len(sender.calls) == 1
    assert "源头断供" in sender.calls[0][1]


def test_scan_no_silence_alert_when_new_files_ingested(tmp_path):
    """有新文件即证明源头活着——此时绝不判断供，无论 ledger 里的旧时间戳多老。"""
    export_dir = tmp_path / "export"
    export_dir.mkdir()
    _make_export_xlsx(export_dir / "old.xlsx", [_row()])
    ledger_path = tmp_path / "ledger.json"
    conn = _FakeFullConnector({"26327000000742719331": [{"DocNo": "AP-1", "InvoiceNo": "42719331"}]},
                              {"AP-1": []})
    scan_once(export_dir, tmp_path / "out", ledger_path, conn, now="2026-07-01T00:00:00+00:00")

    _make_export_xlsx(export_dir / "new.xlsx", [_row(digital_no="26327000000742719999")])
    sender = _RecordingSender()
    outcome = scan_once(export_dir, tmp_path / "out", ledger_path, conn,
                        now="2026-08-17T02:30:00+00:00",
                        webhook_url="https://example.invalid/w", sender=sender)

    assert outcome.result.files_processed == ["new.xlsx"]
    assert outcome.source_silence is None
    assert sender.calls == []


def test_scan_silence_without_webhook_still_detects(tmp_path):
    """webhook 未配置时仍如实判定（只是发不出去）——当前 `.51` 真实状态即此。"""
    export_dir = tmp_path / "export"
    export_dir.mkdir()
    _make_export_xlsx(export_dir / "one.xlsx", [_row()])
    ledger_path = tmp_path / "ledger.json"
    conn = _FakeFullConnector({"26327000000742719331": [{"DocNo": "AP-1", "InvoiceNo": "42719331"}]},
                              {"AP-1": []})
    scan_once(export_dir, tmp_path / "out", ledger_path, conn, now="2026-08-10T13:11:46+00:00")

    outcome = scan_once(export_dir, tmp_path / "out", ledger_path, conn,
                        now="2026-08-17T02:30:00+00:00", webhook_url=None)

    assert outcome.source_silence is not None
    assert outcome.alert_sent is False
    assert outcome.alert_error == ""


def test_file_level_failure_takes_priority_over_silence(tmp_path):
    """两类告警互斥、文件级失败优先——不同时发两条。"""
    export_dir = tmp_path / "export"
    export_dir.mkdir()
    ledger_path = tmp_path / "ledger.json"
    conn = _FakeFullConnector({"26327000000742719331": [{"DocNo": "AP-1", "InvoiceNo": "42719331"}]},
                              {"AP-1": []})
    _make_export_xlsx(export_dir / "ok.xlsx", [_row()])
    scan_once(export_dir, tmp_path / "out", ledger_path, conn, now="2026-08-10T13:11:46+00:00")

    # 新文件但 sheet 名不符 → 文件级失败（源头其实活着，只是这份解析不了）
    _make_export_xlsx(export_dir / "bad.xlsx", [_row()], sheet_name="完全不相干的表")
    sender = _RecordingSender()
    outcome = scan_once(export_dir, tmp_path / "out", ledger_path, conn,
                        now="2026-08-17T02:30:00+00:00",
                        webhook_url="https://example.invalid/w", sender=sender)

    assert outcome.file_level_failures
    assert outcome.source_silence is None
    assert len(sender.calls) == 1
    assert "文件级摄取失败" in sender.calls[0][1]


def test_build_silence_message_states_facts_without_prescribing():
    from fi2.tax_export_scan import SourceSilence
    msg = build_silence_message(
        export_dir="D:\airead",
        silence=SourceSilence(last_ingest_at="2026-08-10T13:11:46+00:00",
                              workdays_silent=5, threshold=3))
    assert "源头断供" in msg
    assert "5 个工作日" in msg
    assert "D:\airead" in msg
    assert "扫描机制本身运行正常" in msg      # 不让读者误以为是我方坏了
