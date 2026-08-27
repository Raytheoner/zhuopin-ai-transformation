"""税务导出发票明细 Excel 接入（design：openspec/changes/fi2-tax-export-ingest，队列 #295）。

背景：唐燕萍（财务 AI 专员）2026-08-04 拍板 OCR 方案作废，改为「税务系统导出发票明细
Excel → 放固定目录 → 定期提取字段」；2026-08-06 把落盘目录/导出责任人时点/新增文件
判据/完整性校验口径四个决策点全部定死（详见场景 CLAUDE.md 队列 #295 段）。

本模块只产出既有 `feed_source.FeedSource(invoice_sample_dir=...)` 通道已经认识的
`invoice.csv`（字段：inv_no/ap_no/item_code/unit/unit_price/inv_qty/untaxed_amount/
tax_rate/tax_amount/inv_date）——`feed_source.py`/`parse_invoice`/`partition_invoices`
本模块一律不改、原样复用。

两处反查（design D1/D2，均基于 2026-08-07 真实探测，非纸面假设）：

  · ap_no：导出 Excel 无我方内部单号，唯一可用标识是「数电发票号码」。真实探测证实
    `AP/Query.InvoiceNo` 字段**与数电发票号码字面不总是相等**（多数只存后 8 位，少数
    存全串）——改用「后 8 位查询 + 客户端 suffix 校验」，已用全部 8 个真实样本验证
    8/8 唯一命中。`AP/Query` 的 `invoiceNo` 过滤参数服务端是 CONTAINS 语义（非精确
    匹配，用哨兵值验证过），故服务端返回集合不可直接信任，必须客户端二次校验。

  · item_code：导出 Excel 无我方料号（只有「货物或应税劳务名称」+「规格型号」文本，
    此前 65.8% 匹配率已证伪单纯文本匹配路子）。ap_no 已确定后，用该 AP 单下的行项目
    （`ItemCode` 即我方真实料号）按 (数量, 含税单价) 唯一匹配来赋值；命中 0 或 ≥2 行
    时不得猜测，标记未解析、留痕待人工核对（fail-loud，不静默丢弃也不静默猜错）。

发票级幂等（2026-08-24 补，队列 #371）
────────────────────────────────────
**真实生产错误**：唐燕萍 2026-08-21 举证 `AP-2026070036·行10`／`R01F.0034` 面板显示
数量 42,000／未税 145,320／税额 18,891.60，恰为真实发票（`26327000000742719331`）
21,000／72,660／9,445.80 的 **2 倍**，面板据此报「数量金额不符」并 BLOCK 退回。

**根因**：原去重只有**文件内容 SHA256** 一层（`_hash_file`→`discover_new_files`→
`is_processed`）。同一张发票只要出现在两份**字节不同**的 xlsx 里（重导一次、导出
区间重叠、另存一次都会改变字节），两份都会被摄取，`write_invoice_csv` 又是追加写
⇒ 同一发票的行进 `invoice.csv` 两次，面板按 `(ap_no, item_code)` 聚合求和后翻倍。
2026-08-24 在 `.51` 实测坐实：3409 行中 `(inv_no, ap_no, item_code)` 去重后仅 2703，
上述目标发票在 `invoice.csv` 里正好 2 行且逐字段相同，而它在现存源文件里只有 1 行
（另一份来自已被删除的文件）。

**修法＝发票级幂等，「首个文件胜出」**：数电发票号码唯一标识一张发票，发票一经开出
其明细不可变（要改只能红冲重开、换号）⇒ **某个 `inv_no` 已由某个文件贡献过，则此后
任何**其它**文件里的同号发票行一律跳过**。

  🔴 **判据只对「跨文件」成立，同一文件内不去重** —— 真实数据里同一张发票出现同
  料品、同数量、同单价的多行是合法的（实测 `26942000000588188581` 单张发票 60 行、
  其中一组签名重复 6 次，全部来自同一份源文件）。**按行内容去重会误删这些合法行**，
  故本模块**不看行内容**，只看「这张发票是不是别的文件已经贡献过了」。

  🔴 **零新增载体** —— 已摄取发票号集合**直接从 `invoice.csv` 现读**
  （`load_ingested_invoice_nos`），不另立状态文件。理由同 `detect_source_silence`
  复用 ledger：多一份状态就多一处会与真相分叉的地方；而 `invoice.csv` 本身就是
  「哪些发票已经进来了」的唯一真相，重建它即自动重建这道闸，不会出现「CSV 已重建
  但闸还记得旧发票、那批行再也回不来」的锁死态。

  跳过的行**不进 `diagnostics`** —— 那是「未解析、需人工核对」的留痕，而跨文件重复
  是预期内的正常现象（她的导出区间本就会重叠），混进去会淹没真正的失败信号，也会
  被 `tax_export_scan` 的文件级失败判定误读。改为独立计数
  （`duplicate_rows_skipped`／`duplicate_invoice_nos`）由调用方如实打印——**不静默**。

未解析行必须可重试（2026-08-26 补，队列 #418）
──────────────────────────────────────────────
**真实生产错误**：唐燕萍 2026-08-26 随机抽 10 张 AP 单复核，**4 张被面板报「无发票
支撑」而发票实际存在**（她自带四组对照，见队列 #418）。

**根因＝摄取期的「一次性判决」**：本模块此前对每一行只有一次机会 ——

  ⑴ `resolve_ap_no` 零命中/歧义、或 `resolve_item_code` 对不上 ⇒ 该行**不进
     `resolved_rows`**，只落一条 `IngestDiagnostic`；
  ⑵ 而 `IngestDiagnostic` **只被 CLI 打印到 stdout，从不落盘**；
  ⑶ 与此同时该文件的 SHA256 已被 `mark_processed` 写进 `.processed_exports.json`
     ⇒ 此后 `discover_new_files` 永远跳过它，**那些行再也不会被重试**。

  🔑 **为什么这必然发生，而不是偶发**：`resolve_ap_no` 反查的是 U9C 里**当时**存在的
  AP 单。发票开出与应付单立账之间天然有时间差 —— **一张 4 月开出的发票，其 AP 单
  可能 8 月才立**。摄取跑在两者之间，零命中就是必然结果，而不是数据脏。她那四组里
  有两组的发票正落在《…20260401-20260430》导出文件里、AP 单却是 8 月的，形态吻合。

  🔴 **这条缺陷的危险不在它报错，而在它报的是一个看起来完全合理的结论** ——
  「无发票支撑（发票还没到）」在账上天天真实发生，因此没人会怀疑那批里混着假的。
  **错误不产生任何信号。**

**修法＝未解析行落进 ledger，每次运行先重试**：

  ① **登记**：`ap_no_zero_match`／`ap_no_ambiguous`／`item_code_zero_match`／
     `item_code_ambiguous` 四类未解析行，连同 `row_index`／`数电发票号码` 一并写进该
     文件的 ledger 条目 `unresolved`（**只登记「U9C 侧状态变了就可能解开」的四类**；
     `数电发票号码` 本身为空这类永不会自愈，仍只作诊断，不进重试队列）。
  ② **重试**：每次 `ingest_directory` 先跑一遍重试 pass —— 对 ledger 里仍有
     `unresolved`、且**源文件仍在盘上且 SHA256 与 ledger 键一致**的文件，重新解析并
     只重跑那几行。解到了就并进 `resolved_rows` 并从 `unresolved` 移除；仍解不开就
     原样留着，下次接着试。
  ③ 🔴 **重试不了的必须出声**：源文件已被删除/已改动（SHA 不符）⇒ 那些行**永远
     不可能再解开**，计入 `unretryable_unresolved`／`unretryable_files` 由调用方如实
     打印。**不静默**——这正是本缺陷当初得以潜伏的方式。

  **零新增载体**（同 #371 口径）：重试队列直接寄在既有 `.processed_exports.json` 的
  条目里，不另立状态文件。

  ⚠️ **与 `scripts/rebuild_invoice_csv.py` 闸①/闸② 的相互作用（必须一并改，否则会
  把那把尺子弄坏）**：该脚本按「ledger 每份文件贡献一个**连续块**」把 `invoice.csv`
  切回源文件。重试会让**同一份文件在 CSV 尾部再追加一段**，其贡献不再连续 ⇒ 旧的
  「按 (processed_at, 文件名) 排序后按 row_count 累加」模型当场失真。故 ledger 条目
  改为记 `segments`（`[{"seq": 递增序号, "row_count": n}, ...]`），`row_count` 仍是
  各段之和（闸① 口径不变）；`partition_by_ledger` 改为按 `seq` 展平所有段。老 ledger
  没有 `segments` 时按原 (processed_at, 文件名) 次序一次性补齐，行为等价。
"""
from __future__ import annotations

import csv
import hashlib
import json
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import openpyxl

_SHEET_NAME = "信息汇总表"

# 「信息汇总表」sheet 必需列（缺任一列即判定该文件解析失败，不产出部分数据）。
_REQUIRED_COLUMNS = (
    "数电发票号码", "货物或应税劳务名称", "规格型号", "单位",
    "数量", "单价", "金额", "税率", "税额", "开票日期",
)

_INVOICE_CSV_FIELDS = [
    "inv_no", "ap_no", "item_code", "unit", "unit_price",
    "inv_qty", "untaxed_amount", "tax_rate", "tax_amount", "inv_date",
]

# item_code 反查时 (数量, 含税单价) 的相对容差——非 config.py R1-R7 业务容差体系
# 的一部分，纯粹用于摄取阶段的行匹配去噪（真实浮点误差量级，见 design D2）。
_QTY_REL_TOL = 1e-6
_PRICE_REL_TOL = 1e-4

# (未税金额, 税额) 反查口径下的**绝对**容差（元）。用绝对值而非相对值，是因为这两个
# 数在税务导出与 U9C 两侧都已是「结算到分」的终值，真正的噪声只有分位进位，与金额
# 大小无关；用相对容差反而会让大额行的容忍度大到失真。
_AMOUNT_ABS_TOL = 0.01

# 组合匹配（⒞）时允许参与求和的 AP 行数上限。超过即不尝试——组合数随行数指数增长，
# 而真实数据里「发票按合计开票」拆的是个位数行，放大上限只会换来指数级耗时。
_SUBSET_SUM_MAX_ROWS = 16

# ── item_code 反查的匹配口径（队列 #424）───────────────────────────────────
#
# 🔴 **本行是唯一开关；口径未经唐燕萍签认前不得改动。**
# 现值 `"qty_price"` ＝ 2026-08-07 上线至今一字未改的行为，本次修改**不默认生效**
# （判据/口径类永不默认生效，CLAUDE.md §5 IATF 显式签认红线）。
#
# ⚠️ 为什么开关必须是一个模块级常量、而不是环境变量或表单选项：本项目 `.51` 的部署
# 方式是「整包同步」（队列 #418 ⑻ 实测坐实），任何一次别的变更包部署都会把 master
# 上的这份代码一并带上生产 ⇒ **「合入 master 但留步不生效」只能靠默认值本身守住**，
# 不能靠「先别部署」。改这一行等于改生产口径，故它必须是一次显式的、可 grep 的提交。
#
# 各候选的语义与实测影响面见 `scripts/probe_424_itemcode_candidates.py` 的产出。
_ITEM_MATCH_STRATEGY = "qty_price"

#: 全部已实现的候选口径（`resolve_item_code(strategy=...)` 的合法取值）。
#: 🔴 逐条对照队列 #424 的 ⒜⒝⒞ —— 命名刻意写全，不用 a/b/c，避免下游读到缩写还要回查。
ITEM_MATCH_STRATEGIES: tuple[str, ...] = (
    "qty_price",                    # 现状：(数量, 含税单价) 逐行唯一匹配
    "amount",                       # ⒜：改用 (未税金额, 税额) 逐行唯一匹配
    "qty_price_then_amount",        # ⒜ 的保守形态：现状优先，零命中再试金额
    "qty_price_then_single_item",   # ⒝ 的可落地形态：现状优先，零命中且该 AP 单只有一个料号 ⇒ 取之
    "qty_price_then_subset_sum",    # ⒞：现状优先，零命中再试「若干 AP 行数量合计 == 发票数量」
)


@dataclass
class IngestDiagnostic:
    """一条无法自动解析/摄取的记录（不静默丢弃，见 spec「未解析记录留痕不静默丢弃」）。"""
    file: str
    reason: str          # "sheet_missing" / "field_missing" / "ap_no_zero_match" /
                          # "ap_no_ambiguous" / "item_code_zero_match" / "item_code_ambiguous"
    detail: str = ""
    digital_invoice_no: Optional[str] = None
    row_index: Optional[int] = None   # 1-based，sheet 内数据行号；文件级问题为 None


@dataclass
class IngestResult:
    resolved_rows: list[dict] = field(default_factory=list)     # 已就绪的 invoice.csv 行
    diagnostics: list[IngestDiagnostic] = field(default_factory=list)
    files_processed: list[str] = field(default_factory=list)
    files_skipped: list[str] = field(default_factory=list)      # 已在已处理清单中，本次跳过
    # 发票级幂等（队列 #371）：因该发票号已由**别的**文件贡献过而跳过的行。
    # 刻意不进 `diagnostics`——那是「需人工核对」的留痕，跨文件重复属预期正常现象。
    duplicate_rows_skipped: int = 0
    duplicate_invoice_nos: list[str] = field(default_factory=list)
    # ── 未解析行重试（队列 #418）────────────────────────────────────────────
    # 本次重试 pass 里终于解开的行（此前批次报「无发票支撑」的假报，就是它们）。
    retried_rows_resolved: int = 0
    retried_invoice_nos: list[str] = field(default_factory=list)
    # 跑完后仍未解开、已留在 ledger 里等下次重试的行数（正常态，非故障）。
    pending_unresolved: int = 0
    # 🔴 源文件已不在盘上或已改动（SHA 与 ledger 键不符）⇒ 这些行**永远解不开了**。
    # 必须由调用方如实打印——不静默（见模块 docstring ③）。
    unretryable_unresolved: int = 0
    unretryable_files: list[str] = field(default_factory=list)


# ── 已处理清单（内容哈希，design D4）───────────────────────────────────────

def _hash_file(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def load_ledger(ledger_path: Path | str) -> dict:
    p = Path(ledger_path)
    if not p.exists():
        return {}
    return json.loads(p.read_text(encoding="utf-8"))


def save_ledger(ledger_path: Path | str, ledger: dict) -> None:
    p = Path(ledger_path)
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(json.dumps(ledger, ensure_ascii=False, indent=2), encoding="utf-8")


def is_processed(ledger: dict, file_hash: str) -> bool:
    return file_hash in ledger


# ── 发票级幂等闸（队列 #371）──────────────────────────────────────────────

def load_ingested_invoice_nos(invoice_csv_path: Path | str) -> set[str]:
    """读出 `invoice.csv` 里已有的全部 `inv_no`——即「哪些发票已经进来了」。

    **零新增载体**（见模块 docstring）：不另立状态文件，`invoice.csv` 自己就是唯一
    真相。文件不存在（首次摄取）返回空集合，**不报错**——「还没开始」不是异常。

    ⚠️ 表头缺失/字段名不符时同样返回空集合而非抛错：本函数只是一道去重闸，读不到
    就退化为「不去重」＝原有行为，绝不因为闸本身读不出而让整次摄取失败。
    """
    p = Path(invoice_csv_path)
    if not p.exists():
        return set()
    with open(p, encoding="utf-8-sig", newline="") as f:
        return {
            str(row["inv_no"]).strip()
            for row in csv.DictReader(f)
            if row.get("inv_no")
        }


# ── 贡献分段与重试队列（队列 #418）────────────────────────────────────────
#
# 条目形状（`.processed_exports.json`，键＝源文件内容 SHA256）：
#   {"file": "xxx.xlsx", "row_count": 12, "processed_at": "...",
#    "segments":   [{"seq": 3, "row_count": 12}],          # 该文件贡献的各连续段
#    "unresolved": [{"row_index": 5, "digital_invoice_no": "263...",
#                    "reason": "ap_no_zero_match", "detail": ""}]}
#
# `row_count` 恒等于各 `segments[].row_count` 之和（`rebuild_invoice_csv` 闸① 口径
# 不变）；`seq` 是**全 ledger 范围**的追加序号，等于该段在 `invoice.csv` 里的先后。

# 只有这四类未解析原因值得重试——它们全都取决于 U9C 侧**当时**的状态（AP 单还没立、
# 行项目还没改对），状态一变就可能自行解开。`digital_invoice_no_missing`（发票号本身
# 为空）永远不会自愈，故不进重试队列，仍只作诊断。
_RETRYABLE_REASONS = frozenset({
    "ap_no_zero_match", "ap_no_ambiguous", "item_code_zero_match", "item_code_ambiguous",
})


def _entry_segments(entry: dict) -> list[dict]:
    return entry.get("segments") or []


def ensure_segments(ledger: dict) -> None:
    """给老 ledger 条目一次性补齐 `segments`（队列 #418 引入本字段前写的条目）。

    补齐次序＝`rebuild_invoice_csv.partition_by_ledger` 此前使用的
    `(processed_at, file)` —— 即那些条目当初真实的追加次序，故补齐前后切分结果等价。
    幂等：已有 `segments` 的条目不动。
    """
    legacy = sorted(
        (kv for kv in ledger.items() if not _entry_segments(kv[1])),
        key=lambda kv: (kv[1].get("processed_at", ""), kv[1].get("file", "")),
    )
    if not legacy:
        return
    # 已有 segments 的条目占掉的 seq 必须避开（混合态：一部分条目已迁移过）。
    used = {s.get("seq", 0) for _h, v in ledger.items() for s in _entry_segments(v)}
    seq = 0
    for _h, entry in legacy:
        while seq in used:
            seq += 1
        entry["segments"] = [{"seq": seq, "row_count": int(entry.get("row_count", 0))}]
        used.add(seq)
        seq += 1


def next_seq(ledger: dict) -> int:
    """下一个可用的全局追加序号。"""
    seqs = [s.get("seq", -1) for _h, v in ledger.items() for s in _entry_segments(v)]
    return max(seqs, default=-1) + 1


def mark_processed(ledger: dict, file_hash: str, filename: str, *, row_count: int,
                   processed_at: str, seq: int | None = None,
                   unresolved: list[dict] | None = None) -> None:
    """登记一份**新**文件的处理结果。

    `seq`：本次贡献段在 `invoice.csv` 里的追加序号（`None` 时自动取 `next_seq`）。
    `unresolved`：本次未解开、留待后续重试的行（队列 #418），空列表即「全部解开了」。
    """
    ledger[file_hash] = {
        "file": filename,
        "row_count": row_count,
        "processed_at": processed_at,
        "segments": [{"seq": next_seq(ledger) if seq is None else seq, "row_count": row_count}],
        "unresolved": list(unresolved or ()),
    }


def append_segment(entry: dict, *, seq: int, row_count: int, processed_at: str) -> None:
    """给**已处理过**的文件追加一段贡献（队列 #418 重试 pass 解开了它此前未解的行）。

    `row_count` 同步累加，`rebuild_invoice_csv` 闸①（各条目 row_count 之和 == CSV
    行数）因此仍然成立；而 `segments` 保住了「哪一段在 CSV 的哪个位置」，闸②/闸③
    的归属重建不会因为一份文件贡献不连续而失真。
    """
    if row_count <= 0:
        return
    entry.setdefault("segments", []).append({"seq": seq, "row_count": row_count})
    entry["row_count"] = int(entry.get("row_count", 0)) + row_count
    entry["processed_at"] = processed_at


# ── Excel 解析 ────────────────────────────────────────────────────────────

def _resolve_sheet_name(sheetnames: list[str]) -> Optional[str]:
    """匹配「信息汇总表」sheet：精确名优先，否则退而取第一个以该名为前缀的 sheet。

    队列 #82 第二班拆件巡逻真实数据比对发现：唐燕萍手工导出的真实批量文件里，该
    sheet 实际命名为「信息汇总表1」（round-1 那 8 个验证样本里是精确的「信息汇总表」，
    无后缀）——两者列结构逐字段核对完全一致，只是 sheet 名多一个后缀，故按前缀匹配
    兼容两种命名，不改变列结构解析逻辑。
    """
    if _SHEET_NAME in sheetnames:
        return _SHEET_NAME
    for name in sheetnames:
        if name.startswith(_SHEET_NAME):
            return name
    return None


def parse_export_workbook(path: Path | str) -> list[dict]:
    """解析「信息汇总表」（或「信息汇总表1」等前缀变体）sheet，返回逐行原始字典（键=表头原文）。

    工作表缺失/必需列缺失时抛 ValueError（调用方据此产出 sheet_missing/field_missing
    诊断，不产出该文件的任何发票明细记录——spec「工作表缺失或结构不符」场景）。
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet_name = _resolve_sheet_name(wb.sheetnames)
    if sheet_name is None:
        raise ValueError(f"缺少「{_SHEET_NAME}」工作表（现有 sheet：{wb.sheetnames}）")
    ws = wb[sheet_name]
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    header = [str(h).strip() if h is not None else "" for h in header_row]
    missing = [c for c in _REQUIRED_COLUMNS if c not in header]
    if missing:
        raise ValueError(f"「{_SHEET_NAME}」缺少必需列：{missing}")

    rows: list[dict] = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        if raw is None or all(v is None for v in raw):
            continue
        rows.append(dict(zip(header, raw)))
    return rows


def _parse_tax_rate(v) -> float:
    """「13%」这类字符串 → 0.13；已是数值则原样转 float。"""
    if v is None:
        return 0.0
    s = str(v).strip()
    if s.endswith("%"):
        return float(s[:-1]) / 100.0
    return float(s)


def _parse_date(v) -> str:
    """「2026-07-10 11:33:22」→「2026-07-10」（同既有 `_u9c_date` 截取惯例）。"""
    if v is None:
        return ""
    s = str(v).strip()
    return s[:10]


# ── ap_no 反查（design D1）───────────────────────────────────────────────

def resolve_ap_no(connector, digital_invoice_no: str) -> tuple[Optional[str], str, str]:
    """用「数电发票号码」反查唯一 ap_no。

    返回 (ap_no_or_None, reason, detail)：
      · 唯一命中 → (ap_no, "", "")
      · 零命中   → (None, "ap_no_zero_match", "")
      · 歧义     → (None, "ap_no_ambiguous", "候选：<ap_no 列表>")

    不假设服务端 `invoiceNo` 过滤是精确匹配（2026-08-07 真实探测证实是 CONTAINS
    语义）——用后 8 位查询后，仍对每一候选行的 `InvoiceNo` 做客户端 suffix 校验，
    只有 `digital_invoice_no.endswith(候选.InvoiceNo)` 才采信。
    """
    suffix = digital_invoice_no[-8:] if len(digital_invoice_no) > 8 else digital_invoice_no
    candidates = connector.get_ap_lines_by_invoice_no(suffix)
    ap_nos: set[str] = set()
    for row in candidates:
        stored = row.get("InvoiceNo")
        if stored and digital_invoice_no.endswith(str(stored)):
            doc_no = row.get("DocNo")
            if doc_no:
                ap_nos.add(str(doc_no))
    if not ap_nos:
        return None, "ap_no_zero_match", ""
    if len(ap_nos) > 1:
        return None, "ap_no_ambiguous", f"候选：{sorted(ap_nos)}"
    return next(iter(ap_nos)), "", ""


# ── item_code 反查（design D2）───────────────────────────────────────────

def _f(v) -> Optional[float]:
    """原始行字段 → float；`None`/空/非数一律返回 `None`（**不当 0 用**）。

    🔴 缺字段与「值是 0」必须分开：把缺字段读成 0 会让一条根本没有金额的 AP 行去和
    一张金额为 0 的发票行「匹配上」——一个不报错的错。
    """
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _match_by_qty_price(
    ap_rows: list[dict], *, qty: float, taxed_unit_price: float,
    qty_rel_tol: float, price_rel_tol: float,
) -> set[str]:
    """现状口径：(数量, 含税单价) 逐行比，两者都落在相对容差内即算命中。"""
    out: set[str] = set()
    for row in ap_rows:
        ap_qty = _f(row.get("APQtyTU"))
        ap_price = _f(row.get("TaxPrice"))
        item_code = row.get("ItemCode")
        if ap_qty is None or ap_price is None or not item_code:
            continue
        qty_close = abs(ap_qty - qty) <= max(abs(qty), abs(ap_qty)) * qty_rel_tol + 1e-9
        price_close = (abs(ap_price - taxed_unit_price)
                       <= max(abs(taxed_unit_price), abs(ap_price)) * price_rel_tol + 1e-9)
        if qty_close and price_close:
            out.add(str(item_code))
    return out


def _match_by_amount(
    ap_rows: list[dict], *, untaxed_amount: Optional[float], tax_amount: Optional[float],
    abs_tol: float,
) -> set[str]:
    """⒜ 口径：(未税金额, 税额) 逐行比，绝对容差到分。

    🔴 **对计量单位换算与「按包装件 vs 按容量」免疫** —— 队列 #424 实测的
    `AP-2026080041` 密封胶正是此形态：数量差 310 倍（20 支 × 310ML ＝ 6200），
    而未税 655.12／税额 85.16 两侧一分不差。

    ⚠️ 但它**不免疫「发票按合计开票、AP 按行拆分」** —— 那一类里没有任何单独一行
    AP 的金额等于发票行金额（`AP-2026080137` 气泡袋即是），本口径同样零命中。
    发票或 AP 缺这两个字段时返回空集（不猜、不回落到别的键）。
    """
    if untaxed_amount is None or tax_amount is None:
        return set()
    out: set[str] = set()
    for row in ap_rows:
        ap_untaxed = _f(row.get("NonTaxAmtTC"))
        ap_tax = _f(row.get("TaxAmtTC"))
        item_code = row.get("ItemCode")
        if ap_untaxed is None or ap_tax is None or not item_code:
            continue
        if (abs(ap_untaxed - untaxed_amount) <= abs_tol
                and abs(ap_tax - tax_amount) <= abs_tol):
            out.add(str(item_code))
    return out


def _single_item_code(ap_rows: list[dict]) -> set[str]:
    """⒝ 的可落地形态：该 AP 单**通篇只有一个料号**时，这张发票只可能挂在它上面。

    ⚠️ 这不是「AP 单级金额归集」的完整实现，而是它唯一能唯一确定 `item_code` 的那个
    子情形。队列 #424 ⒝ 原文写的是「回落到 AP 单级归集（该发票整体 vs 该 AP 单全部行
    的金额合计）」——**单级归集能判断「这张发票属不属于这张单」，却判断不出「属于哪一
    行料品」**；而 `invoice.csv` 的每一行必须带一个 `item_code`。多料号单据下 ⒝ 无法
    落地成一个行级答案，除非同时接受「把一张发票行拆成多行写入」（见
    `_match_by_subset_sum`）。这一点在选 ⒝ 之前必须先说清。
    """
    codes = {str(r.get("ItemCode")) for r in ap_rows if r.get("ItemCode")}
    return codes if len(codes) == 1 else set()


def _match_by_subset_sum(
    ap_rows: list[dict], *, qty: float, taxed_unit_price: float, price_rel_tol: float,
) -> set[str]:
    """⒞ 口径：允许「若干 AP 行的数量合计 == 发票行数量」（同一含税单价内）。

    队列 #424 实测的 `AP-2026080137` 气泡袋即此形态：发票 6000 ＝ AP 的 1000 ＋ 5000，
    发票 7000 ＝ 4000 ＋ 3000，**加起来分毫不差，但没有任何单独一行等于 6000/7000**。

    🔴 **返回的可能是一个多元素集合** —— 6000 那一笔跨 `J02E.0024` 与 `R02E.0024`
    两个料号。在现行输出契约（一行发票 → 一个 `item_code`）下，它会被上层如实判为
    `item_code_ambiguous`，**而不是被随便挑一个**。要让 ⒞ 真正解开这一类，必须同时
    接受「把一张发票行按 AP 行拆成多行写入 `invoice.csv`」——那是另一个口径决定，
    不在本函数范围内。本函数只负责把「到底能不能凑出来」这件事量清楚。

    先按含税单价分组，再在组内做子集和；组内行数超过 `_SUBSET_SUM_MAX_ROWS` 即放弃
    该组（不静默地退化成部分搜索，直接不试）。
    """
    groups: dict[float, list[tuple[float, str]]] = {}
    for row in ap_rows:
        ap_qty = _f(row.get("APQtyTU"))
        ap_price = _f(row.get("TaxPrice"))
        item_code = row.get("ItemCode")
        if ap_qty is None or ap_price is None or not item_code:
            continue
        if abs(ap_price - taxed_unit_price) > (
                max(abs(taxed_unit_price), abs(ap_price)) * price_rel_tol + 1e-9):
            continue
        groups.setdefault(round(ap_price, 6), []).append((ap_qty, str(item_code)))

    out: set[str] = set()
    for rows in groups.values():
        if len(rows) > _SUBSET_SUM_MAX_ROWS:
            continue
        n = len(rows)
        for mask in range(1, 1 << n):
            total = 0.0
            for i in range(n):
                if mask >> i & 1:
                    total += rows[i][0]
            if abs(total - qty) <= max(abs(qty), abs(total)) * _QTY_REL_TOL + 1e-9:
                out |= {rows[i][1] for i in range(n) if mask >> i & 1}
    return out


def resolve_item_code(
    ap_lines_for_ap_no: list[dict], qty: float, untaxed_unit_price: float, tax_rate: float,
    *,
    untaxed_amount: Optional[float] = None, tax_amount: Optional[float] = None,
    strategy: Optional[str] = None,
    qty_rel_tol: Optional[float] = None, price_rel_tol: Optional[float] = None,
    amount_abs_tol: Optional[float] = None,
) -> tuple[Optional[str], str, str]:
    """在已知 ap_no 的行项目中唯一匹配料品编码。

    `ap_lines_for_ap_no`：该 ap_no 下 `AP/Query` 原始行列表（调用方按 ap_no 取好，
    本函数不发起网络调用，便于测试与复用同一 ap_no 下多张发票行共享一次拉取）。

    返回 (item_code_or_None, reason, detail)，reason 语义同 `resolve_ap_no`。

    `strategy`（队列 #424，默认 `_ITEM_MATCH_STRATEGY` ＝ `"qty_price"` ＝ 现状）：
    见 `ITEM_MATCH_STRATEGIES`。🔴 **口径待唐燕萍签认，默认值不得在签认前改动。**

    `untaxed_amount`/`tax_amount`：`"amount"` 系口径必需（税务导出的「金额」「税额」
    两列）。不传即该口径退化为零命中——**不猜、不由单价反推**（反推会把导出侧的
    四舍五入误差放大成假匹配）。

    三个容差参数留出显式入口，供 `scripts/probe_424_itemcode_candidates.py` 在真实
    全量数据上量「放宽到什么程度分别能捞回多少」，**不供生产调用方随手传**。
    """
    strategy = strategy or _ITEM_MATCH_STRATEGY
    if strategy not in ITEM_MATCH_STRATEGIES:
        raise ValueError(f"未知的 item_code 匹配口径：{strategy!r}；"
                         f"合法值＝{list(ITEM_MATCH_STRATEGIES)}")
    qty_rel_tol = _QTY_REL_TOL if qty_rel_tol is None else qty_rel_tol
    price_rel_tol = _PRICE_REL_TOL if price_rel_tol is None else price_rel_tol
    amount_abs_tol = _AMOUNT_ABS_TOL if amount_abs_tol is None else amount_abs_tol
    taxed_unit_price = untaxed_unit_price * (1 + tax_rate)

    def _by_qty_price() -> set[str]:
        return _match_by_qty_price(
            ap_lines_for_ap_no, qty=qty, taxed_unit_price=taxed_unit_price,
            qty_rel_tol=qty_rel_tol, price_rel_tol=price_rel_tol)

    def _by_amount() -> set[str]:
        return _match_by_amount(
            ap_lines_for_ap_no, untaxed_amount=untaxed_amount,
            tax_amount=tax_amount, abs_tol=amount_abs_tol)

    if strategy == "qty_price":
        matched_codes = _by_qty_price()
    elif strategy == "amount":
        matched_codes = _by_amount()
    else:
        # 🔴 「回落」一律只在**零命中**时发生，绝不在**歧义**时发生：现状口径已经找到
        # 多个候选，说明这张发票行本身就分不清挂哪一行，换把尺子只会换一批候选、
        # 不会让它变清楚——那种「换到能出一个答案为止」正是静默猜测。
        matched_codes = _by_qty_price()
        if not matched_codes:
            if strategy == "qty_price_then_amount":
                matched_codes = _by_amount()
            elif strategy == "qty_price_then_single_item":
                matched_codes = _single_item_code(ap_lines_for_ap_no)
            elif strategy == "qty_price_then_subset_sum":
                matched_codes = _match_by_subset_sum(
                    ap_lines_for_ap_no, qty=qty, taxed_unit_price=taxed_unit_price,
                    price_rel_tol=price_rel_tol)

    if not matched_codes:
        return None, "item_code_zero_match", ""
    if len(matched_codes) > 1:
        return None, "item_code_ambiguous", f"候选：{sorted(matched_codes)}"
    return next(iter(matched_codes)), "", ""


def _join_detail(*parts: str) -> str:
    return "；".join(p for p in parts if p)


def _item_match_diagnosis(
    ap_rows: list[dict], *, ap_no: str, qty: float, untaxed_unit_price: float,
    tax_rate: float, untaxed_amount: Optional[float], tax_amount: Optional[float],
) -> str:
    """给一条挂不上料号的发票行加一段**可 grep 的身份**（队列 #424「让丢行可见并可查」）。

    产出形如 `ap=AP-2026080041 ap行数=1 料号数=1 换口径可解=[amount,single_item]`。

    🔴 三条边界，写的时候就定死，免得日后被当成「顺手放宽一下」的入口：
      ⑴ **不含任何金额/数量的原始值** —— FI2 审计口径是金额不落盘（webapp 页脚已明写），
         这条诊断会随 ledger 与诊断文件长期留存，不能成为金额的第二个出口。
      ⑵ **不改变任何摄取结果** —— 它只描述「若换成别的口径会怎样」，当前口径该丢的
         照丢。换口径要由唐燕萍签认，不由这条诊断代劳。
      ⑶ 纯 CPU、零网络（AP 行已在调用方缓存里），组合搜索有 `_SUBSET_SUM_MAX_ROWS` 封顶。
    """
    taxed_unit_price = untaxed_unit_price * (1 + tax_rate)
    codes = {str(r.get("ItemCode")) for r in ap_rows if r.get("ItemCode")}
    solvable: list[str] = []
    if len(_match_by_amount(ap_rows, untaxed_amount=untaxed_amount,
                            tax_amount=tax_amount, abs_tol=_AMOUNT_ABS_TOL)) == 1:
        solvable.append("amount")
    if len(_single_item_code(ap_rows)) == 1:
        solvable.append("single_item")
    subset = _match_by_subset_sum(ap_rows, qty=qty, taxed_unit_price=taxed_unit_price,
                                  price_rel_tol=_PRICE_REL_TOL)
    if len(subset) == 1:
        solvable.append("subset_sum")
    elif len(subset) > 1:
        # 凑得出来，但跨了多个料号 —— 现行「一行发票 → 一个 item_code」的输出契约下
        # 它仍然无解，除非接受把发票行拆开写。**如实标成另一类，不混进「可解」。**
        solvable.append(f"subset_sum_跨{len(subset)}料号")
    return (f"ap={ap_no} ap行数={len(ap_rows)} 料号数={len(codes)} "
            f"换口径可解={solvable or '无'}")


# ── 编排 ──────────────────────────────────────────────────────────────────

def discover_new_files(export_dir: Path | str, ledger: dict) -> list[tuple[Path, str]]:
    """列出 `export_dir` 下未出现在 `ledger` 中的 `.xlsx` 文件，返回 (路径, 内容哈希) 列表。"""
    out: list[tuple[Path, str]] = []
    for p in sorted(Path(export_dir).glob("*.xlsx")):
        h = _hash_file(p)
        if not is_processed(ledger, h):
            out.append((p, h))
    return out


class _RowResolver:
    """把「一行导出明细 → invoice.csv 行」的解析收成一处，供新文件与重试 pass 共用。

    两条路径此前只在 `ingest_directory` 里存在一份（新文件），队列 #418 加重试 pass
    后必须逐字同规则——抽出来是为了保证它们**不会分叉**，不是为了好看。
    """

    def __init__(self, connector, result: IngestResult):
        self._connector = connector
        self._result = result
        # 同一张发票的多个明细行共享同一「数电发票号码」（真实样本已观察到单张发票
        # 182 行的情形，见 sample_8/AP-2026050057）——按 digital_no 缓存 ap_no 反查结果，
        # 避免对同一发票号重复发起真实网络请求（此前无缓存时曾在真实端点上耗时超 2 分钟）。
        self._ap_no_cache: dict[str, tuple[Optional[str], str, str]] = {}
        self._ap_lines_cache: dict[str, list[dict]] = {}

    def resolve(self, file_name: str, idx: int, raw: dict) -> tuple[Optional[dict], Optional[dict]]:
        """→ (已解析的 invoice.csv 行 或 None, 待重试登记项 或 None)。

        诊断一律登记进 `result.diagnostics`（不静默）；其中属 `_RETRYABLE_REASONS`
        的额外返回一个待重试登记项，由调用方写进 ledger。
        """
        digital_no = str(raw.get("数电发票号码") or "").strip()
        if not digital_no:
            self._result.diagnostics.append(IngestDiagnostic(
                file=file_name, row_index=idx, reason="digital_invoice_no_missing"))
            return None, None   # 发票号为空永不自愈，不进重试队列

        if digital_no not in self._ap_no_cache:
            self._ap_no_cache[digital_no] = resolve_ap_no(self._connector, digital_no)
        ap_no, reason, detail = self._ap_no_cache[digital_no]
        if ap_no is None:
            return None, self._fail(file_name, idx, digital_no, reason, detail)

        if ap_no not in self._ap_lines_cache:
            self._ap_lines_cache[ap_no] = self._connector.get_ap_lines(ap_no)

        qty = float(raw.get("数量") or 0)
        unit_price = float(raw.get("单价") or 0)
        tax_rate = _parse_tax_rate(raw.get("税率"))
        untaxed_amount = _f(raw.get("金额"))
        tax_amount = _f(raw.get("税额"))
        ap_rows = self._ap_lines_cache[ap_no]
        item_code, i_reason, i_detail = resolve_item_code(
            ap_rows, qty, unit_price, tax_rate,
            untaxed_amount=untaxed_amount, tax_amount=tax_amount)
        if item_code is None:
            # 队列 #424「让丢行可见并可查」：光记一个 `item_code_zero_match` 无法回答
            # 「这一行为什么挂不上、换把尺子挂不挂得上」，而那正是唯一能把这批丢行
            # 分类的信息。诊断串里**只放形状、不放金额**（FI2 审计口径＝金额不落盘），
            # 且**不改变任何摄取结果**——纯粹是给这条丢行加一个可 grep 的身份。
            i_detail = _join_detail(i_detail, _item_match_diagnosis(
                ap_rows, ap_no=ap_no, qty=qty, untaxed_unit_price=unit_price,
                tax_rate=tax_rate, untaxed_amount=untaxed_amount, tax_amount=tax_amount))
            return None, self._fail(file_name, idx, digital_no, i_reason, i_detail)

        return {
            "inv_no": digital_no,
            "ap_no": ap_no,
            "item_code": item_code,
            "unit": str(raw.get("单位") or ""),
            "unit_price": unit_price,
            "inv_qty": qty,
            "untaxed_amount": float(raw.get("金额") or 0),
            "tax_rate": tax_rate,
            "tax_amount": float(raw.get("税额") or 0),
            "inv_date": _parse_date(raw.get("开票日期")),
        }, None

    def _fail(self, file_name: str, idx: int, digital_no: str, reason: str, detail: str) -> Optional[dict]:
        self._result.diagnostics.append(IngestDiagnostic(
            file=file_name, row_index=idx, digital_invoice_no=digital_no,
            reason=reason, detail=detail))
        if reason not in _RETRYABLE_REASONS:
            return None
        return {"row_index": idx, "digital_invoice_no": digital_no,
                "reason": reason, "detail": detail}

    def forget(self, digital_no: str) -> None:
        """丢掉某发票号的 ap_no 反查缓存——重试 pass 跨 U9C 状态变化时必须重查。"""
        self._ap_no_cache.pop(digital_no, None)


def _retry_unresolved(
    export_dir: Path, ledger: dict, resolver: _RowResolver, result: IngestResult,
    known: set[str], now: str,
) -> list[dict]:
    """重试 pass（队列 #418）：把此前批次未解开、如今可能已能解开的行捞回来。

    只碰「源文件仍在盘上、且内容 SHA256 与 ledger 键一致」的条目 —— 文件被删或被改
    过就再也无从重试，如实计入 `unretryable_*`，**不静默**。

    返回本次新解出的行（已按 `segments` 的 seq 次序排好，调用方按此序追加写盘）。
    """
    out: list[dict] = []
    # `ap_no` 反查缓存里可能留着**本次运行早些时候**刚查出的结果，但重试针对的正是
    # 「U9C 侧状态变了」的行 —— 每个发票号在本轮重试里先失效一次、只失效一次。
    # 🔴 不能每行都 forget：真实数据里单张发票 182 行，那会退化成 182 次真实网络调用
    # （无缓存时曾实测耗时超 2 分钟，见 `_RowResolver` 注释）。
    refreshed: set[str] = set()
    entries = sorted(
        ((h, v) for h, v in ledger.items() if v.get("unresolved")),
        key=lambda kv: min((s.get("seq", 0) for s in _entry_segments(kv[1])), default=0),
    )
    for file_hash, entry in entries:
        pending = list(entry.get("unresolved") or ())
        file_name = entry.get("file", "")
        path = export_dir / file_name
        if not file_name or not path.is_file() or _hash_file(path) != file_hash:
            result.unretryable_unresolved += len(pending)
            if file_name and file_name not in result.unretryable_files:
                result.unretryable_files.append(file_name)
            continue
        try:
            raw_rows = parse_export_workbook(path)
        except ValueError as e:
            result.diagnostics.append(IngestDiagnostic(
                file=file_name, reason="parse_error", detail=str(e)))
            result.unretryable_unresolved += len(pending)
            if file_name not in result.unretryable_files:
                result.unretryable_files.append(file_name)
            continue

        # `parse_export_workbook` 的行序稳定，且文件 SHA 已核对一致 ⇒ row_index 仍指向
        # 同一行；再用 digital_invoice_no 二次校验，不一致即放弃该条（不猜）。
        by_index = {idx: raw for idx, raw in enumerate(raw_rows, start=2)}
        seen_before_this_file = set(known)
        contributed_here: set[str] = set()
        still_pending: list[dict] = []
        resolved_here: list[dict] = []
        for item in pending:
            idx = item.get("row_index")
            raw = by_index.get(idx)
            digital_no = str(item.get("digital_invoice_no") or "")
            if raw is None or str(raw.get("数电发票号码") or "").strip() != digital_no:
                result.unretryable_unresolved += 1
                if file_name not in result.unretryable_files:
                    result.unretryable_files.append(file_name)
                continue
            if digital_no in seen_before_this_file:
                # 这张发票此后已由别的文件贡献过了——本条自然消解，无须再试。
                result.duplicate_rows_skipped += 1
                if digital_no not in result.duplicate_invoice_nos:
                    result.duplicate_invoice_nos.append(digital_no)
                continue
            if digital_no not in refreshed:
                resolver.forget(digital_no)   # U9C 侧状态可能已变，缓存不能复用
                refreshed.add(digital_no)
            row, retry_item = resolver.resolve(file_name, idx, raw)
            if row is None:
                if retry_item is not None:
                    still_pending.append(retry_item)
                else:
                    result.unretryable_unresolved += 1
                    if file_name not in result.unretryable_files:
                        result.unretryable_files.append(file_name)
                continue
            resolved_here.append(row)
            contributed_here.add(digital_no)
            result.retried_rows_resolved += 1
            if digital_no not in result.retried_invoice_nos:
                result.retried_invoice_nos.append(digital_no)

        entry["unresolved"] = still_pending
        result.pending_unresolved += len(still_pending)
        if resolved_here:
            append_segment(entry, seq=next_seq(ledger), row_count=len(resolved_here),
                           processed_at=now)
            out.extend(resolved_here)
        known |= contributed_here
    return out


def ingest_directory(
    export_dir: Path | str, ledger_path: Path | str, connector, *, now: str,
    known_invoice_nos: set[str] | None = None, retry_unresolved: bool = True,
) -> IngestResult:
    """扫描目录 → 跳过已处理 → 解析 → ap_no/item_code 反查 → 产出结果（不落盘，见
    `write_invoice_csv`）。`now` 由调用方传入处理时间戳（ISO 字符串），保持本函数
    纯净可测（不读系统时钟）。

    `known_invoice_nos`＝**此前批次**已摄取过的数电发票号码集合（队列 #371 发票级
    幂等闸，调用方用 `load_ingested_invoice_nos(<out_dir>/invoice.csv)` 取）。命中者
    其行一律跳过并计入 `duplicate_rows_skipped`。**本函数不修改传入的集合**（复制一份
    用），调用方可安全复用。

    ⚠️ **传 None 不等于关掉闸** —— 它只表示「没有历史包袱」，本次运行内**跨文件**的
    重复照样会被挡（本次新处理的每份文件都会把自己贡献的发票号并进闸）。要让闸完全
    生效必须传 `known_invoice_nos`，否则跨批次的重复（第二天的新文件含昨天的发票）
    仍会漏过——**#371 那次翻倍正是跨批次形态**。

    🔴 **闸只挡跨文件重复**：每份文件开始处理前先快照一次「已知发票号」，该文件自身
    新贡献的发票号在**本文件处理完之后**才并入快照 —— 故同一文件内同号发票的多行
    （真实数据里合法且常见）全部保留，不会被自己挡掉。

    `retry_unresolved`（队列 #418，默认开）：先跑一遍重试 pass，把此前批次因
    `ap_no_zero_match` 等四类原因未解开、如今 U9C 侧已能解开的行捞回来（**「发票 4 月
    开、AP 单 8 月立」是常态，不是脏数据**，见模块 docstring）。传 `False` 可关掉，
    仅用于单测隔离新文件路径；生产调用不要关——关掉就退回「一次判决、永不复议」的
    旧行为，也就是 #418 那批假「无发票支撑」的成因。

    🔴 **`resolved_rows` 的次序即写盘次序**：重试段在前、新文件段在后，与 ledger 里
    各 `segments[].seq` 严格一致。调用方必须按原序一次性 `write_invoice_csv`，否则
    `rebuild_invoice_csv` 的归属重建会与实际错位。
    """
    export_dir = Path(export_dir)
    ledger = load_ledger(ledger_path)
    ensure_segments(ledger)     # 老 ledger 一次性补齐 segments（队列 #418）
    result = IngestResult()
    known: set[str] = set(known_invoice_nos or ())
    resolver = _RowResolver(connector, result)

    # ── ① 重试 pass：先捞回此前未解开的行（队列 #418）──────────────────────
    if retry_unresolved:
        result.resolved_rows.extend(
            _retry_unresolved(export_dir, ledger, resolver, result, known, now))

    # ── ② 常规 pass：处理新文件 ────────────────────────────────────────────
    new_files = discover_new_files(export_dir, ledger)
    all_files = sorted(export_dir.glob("*.xlsx"))
    new_paths = {p for p, _ in new_files}
    for p in all_files:
        if p not in new_paths:
            result.files_skipped.append(p.name)

    for path, file_hash in new_files:
        try:
            raw_rows = parse_export_workbook(path)
        except ValueError as e:
            result.diagnostics.append(IngestDiagnostic(file=path.name, reason="parse_error", detail=str(e)))
            continue

        rows_here: list[dict] = []
        unresolved_here: list[dict] = []
        # 本文件开始处理前的快照——闸只挡「别的文件已贡献过」，不挡本文件自身
        # 同号发票的多行（真实数据里合法且常见，见模块 docstring）。
        seen_before_this_file = set(known)
        contributed_here: set[str] = set()
        for idx, raw in enumerate(raw_rows, start=2):
            digital_no = str(raw.get("数电发票号码") or "").strip()
            if digital_no and digital_no in seen_before_this_file:
                # 该发票已由别的文件（或此前批次）贡献过 —— 跳过，不重复计数。
                # 刻意不进 diagnostics（见 IngestResult 字段注释），但如实计数。
                result.duplicate_rows_skipped += 1
                if digital_no not in result.duplicate_invoice_nos:
                    result.duplicate_invoice_nos.append(digital_no)
                continue

            row, retry_item = resolver.resolve(path.name, idx, raw)
            if row is None:
                if retry_item is not None:
                    unresolved_here.append(retry_item)
                continue
            rows_here.append(row)
            contributed_here.add(digital_no)

        # 本文件贡献的发票号在此刻才并入闸——保证同文件内多行不自挡（见 docstring）。
        known |= contributed_here
        result.resolved_rows.extend(rows_here)
        result.pending_unresolved += len(unresolved_here)
        mark_processed(ledger, file_hash, path.name, row_count=len(rows_here),
                       processed_at=now, unresolved=unresolved_here)
        result.files_processed.append(path.name)

    save_ledger(ledger_path, ledger)
    return result


def summarize_diagnostics(result: IngestResult) -> list[tuple[str, int, int]]:
    """按 `reason` 汇总本次未解析记录 → `[(reason, 行数, 涉发票张数), ...]`（降序）。

    逐条打印在真实数据上是 26,000 行起步（队列 #418 ⑺），**没人会读完，也就等于
    没人在看** —— 汇总是让这批丢行第一次变成一个能被看一眼的数。
    """
    rows: dict[str, int] = {}
    invs: dict[str, set[str]] = {}
    for d in result.diagnostics:
        rows[d.reason] = rows.get(d.reason, 0) + 1
        if d.digital_invoice_no:
            invs.setdefault(d.reason, set()).add(d.digital_invoice_no)
    return sorted(
        ((r, n, len(invs.get(r, ()))) for r, n in rows.items()),
        key=lambda t: (-t[1], t[0]),
    )


def write_diagnostics_jsonl(result: IngestResult, path: Path | str, *, now: str) -> int:
    """把本次未解析记录**落盘**成可 grep 的 JSONL，返回写入条数（队列 #424）。

    🔴 **这正是 #418 根因链的第 ⑵ 环** —— 诊断此前只被 CLI 打印到 stdout、从不落盘，
    于是「哪些发票行被丢掉了、为什么」这件事在生产上根本不存在记录；#418 的重试队列
    解决了「还有没有下一次机会」，**没有**解决「已经丢掉的这批看不看得见」。本函数只
    补后者：**追加写，不覆盖**，一行一条，供事后按 `reason`／发票号／AP 单号回查。

    ⚠️ **不含金额与数量的原始值**（同 `_item_match_diagnosis` 边界⑴）：本文件长期留存，
    不能成为财务金额的第二个出口；`detail` 里只有形状与「换口径可解与否」。
    """
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    with open(p, "a", encoding="utf-8") as f:
        for d in result.diagnostics:
            f.write(json.dumps({
                "ts": now, "file": d.file, "row_index": d.row_index,
                "reason": d.reason, "digital_invoice_no": d.digital_invoice_no,
                "detail": d.detail,
            }, ensure_ascii=False) + "\n")
    return len(result.diagnostics)


def write_invoice_csv(rows: list[dict], out_path: Path | str) -> None:
    """把已解析行写入 `invoice.csv`（既有 `invoice_sample_dir` 通道认识的字段格式）。

    追加写入：若目标文件已存在（此前批次摄取过），保留旧行、追加新行。

    🔴 **本函数本身不做去重**（2026-08-24 队列 #371 更正此处口径）。去重现在有**两**
    层，缺一层就会出现「面板数字翻倍」：

      ① **文件层**（`.processed_exports.json`，内容 SHA256）——同一份**字节相同**的
         源文件不会被摄取两次。
      ② **发票层**（`ingest_directory(known_invoice_nos=...)`，队列 #371 新增）——
         同一张**数电发票**不会由两份不同文件各贡献一次。

    2026-08-24 之前只有 ①，而她的导出区间本就会重叠、重导一次即改变字节 ⇒ 同一发票
    经两份文件各入库一次，本函数照单追加，面板聚合后翻倍。**调用方必须传
    `known_invoice_nos`**，否则退化回只有 ① 的旧行为。
    """
    p = Path(out_path)
    p.parent.mkdir(parents=True, exist_ok=True)
    file_exists = p.exists()
    with open(p, "a" if file_exists else "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=_INVOICE_CSV_FIELDS)
        if not file_exists:
            w.writeheader()
        for r in rows:
            w.writerow(r)
