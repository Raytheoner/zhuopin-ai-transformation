"""从工作汇总.xlsx「立项门禁」页导出 82 条规则元数据 → data/rules/registry.json。

规则即数据（design.md D2）：工作汇总.xlsx 为单一可信源，规则注册表是其
机器可读快照。规则集演进 → 重跑本脚本 → 登记规则版本号 → 走黄金基准回归。

用法：
    py scripts/build_rule_registry.py [path-to-工作汇总.xlsx] [rule_version]
默认读 ../../../7-外部文档/质量部/AI质量智能建设就绪工作汇总.xlsx。
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

import openpyxl

HERE = Path(__file__).resolve().parent
SCENE = HERE.parent
DEFAULT_XLSX = SCENE / "../../../7-外部文档/质量部/AI质量智能建设就绪工作汇总.xlsx"
# 收口-1 权威权重/严重度表（2026-07-03，陈忱交付）
DEFAULT_WEIGHT_XLSX = (SCENE / "../../../7-外部文档/质量部/质量部AI建立相关文件（6.29-7-3）/"
                       "开发项目立项门禁扣分权重标准_82条规则.xlsx")
OUT = SCENE / "data/rules/registry.json"

# 「立项门禁」页表头（A1:K1）→ 规范字段名
COLS = {
    "规则ID": "rule_id", "章节": "section", "检查项": "check_item",
    "适用类型": "applies_to", "严重度": "severity", "检查类型": "check_type",
    "判定标准（通过条件）": "pass_condition", "阈值/取值域": "threshold",
    "取数来源": "source_anchor", "自动化": "automation", "备注": "trial_result",
}

# 「立项门禁」J 列严重度（阻断/重要/一般/提示）→「封面」三级（错误/警告/提示），
# 仅用于报告分组展示。**一票否决严格取「阻断」**（收口-1 权威裁定，见下）。
SEVERITY_MAP = {
    "阻断": "错误",
    "重要": "警告",   # 收口-1 定案：重要=0.5 系数扣分，非一票否决
    "一般": "警告",
    "提示": "提示",
}

# 严重度权重系数（收口-1 权威：开发项目立项门禁扣分权重标准_82条规则.xlsx）。
# 单条扣分值 = 模块基础分 × 系数 / 模块内系数之和；仅「阻断」触发一票否决。
SEVERITY_COEFF = {"阻断": 1.0, "重要": 0.5, "一般": 0.3, "提示": 0.1}

# 自动化分类 → 实现类别（design 三分类）
AUTOMATION_CLASS = {
    "可机器核": "A",   # 确定性纯函数
    "半自动": "B",     # LLM 语义判定
    "转人工": "C",     # 只验在否
}

# 14 模块基础分（收口-1 权威：扣分权重标准_82条规则.xlsx「模块权重分配」，合计 100）。
# 模块八分（一）成本效益/（二）收益分析说明，基础分不同。
MODULE_BASE_SCORES = {
    "一": 15, "二": 5, "三": 5, "四": 8, "五": 5, "六": 8, "七": 4,
    "八（一）": 10, "八（二）": 15, "九": 8, "十": 6, "十一": 3, "十二": 5, "十三": 3,
}
# 合格分档阈值（封面 + 权重表一致）
TIER_THRESHOLDS = {"合格": 80, "有条件合格": 60}  # ≥80 合格；60~80 有条件；<60 不合格


def _module_key(section: str) -> str:
    """章节文本 → 模块键（对齐 MODULE_BASE_SCORES）。模块八按（一）/（二）细分。"""
    s = _norm(section)
    m = re.match(r"^(十[一二三]?|[一二三四五六七八九])[、，]", s)
    if not m:
        return ""
    ordn = m.group(1)
    if ordn == "八":
        if "（一）" in s or "(一)" in s:
            return "八（一）"
        if "（二）" in s or "(二)" in s:
            return "八（二）"
        return "八（一）"
    return ordn


def _load_weight_severities(weight_path: Path) -> dict[str, str]:
    """从权重表各模块明细页读 rule_id → 严重度（收口-1 权威，覆盖 6-27 基表旧值）。"""
    wb = openpyxl.load_workbook(weight_path, data_only=True)
    out: dict[str, str] = {}
    for ws in wb.worksheets:
        if not _norm(ws.title).startswith("模块"):
            continue
        for row in ws.iter_rows(values_only=True):
            if not row or not isinstance(row[0], (int, float)):
                continue
            rid = str(int(row[0]))
            sev = _norm(row[3]) if len(row) > 3 else ""
            if sev in SEVERITY_COEFF:
                out[rid] = sev
    return out


def build(xlsx_path: Path, rule_version: str, weight_path: Path | None = None) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["立项门禁"]
    header = [(_norm(c.value)) for c in ws[1]]
    idx = {COLS[h]: i for i, h in enumerate(header) if h in COLS}

    # 收口-1 权威严重度覆盖（7-03 权重表 vs 6-27 基表分歧以权重表为准）
    weight_sev = _load_weight_severities(weight_path) if weight_path and weight_path.exists() else {}
    overrides: list[str] = []

    rules = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None or _norm(row[0]) == "":
            continue
        rec = {key: _norm(row[i]) if i < len(row) else "" for key, i in idx.items()}
        rid = str(rec.get("rule_id", ""))
        base_sev = rec.get("severity", "")
        sev = weight_sev.get(rid, base_sev)
        if rid in weight_sev and weight_sev[rid] != base_sev:
            overrides.append(f"规则{rid}: {base_sev}→{sev}")
        auto = rec.get("automation", "")
        rec["severity"] = sev                    # 以权威表为准
        rec["severity_source_6_27"] = base_sev    # 保留基表原值，可审计
        rec["severity_level"] = SEVERITY_MAP.get(sev, "警告")
        rec["impl_class"] = AUTOMATION_CLASS.get(auto, "A")
        rec["coefficient"] = SEVERITY_COEFF.get(sev, 0.3)
        rec["blocking"] = sev == "阻断"          # 收口-1：仅阻断一票否决
        rec["module_key"] = _module_key(rec.get("section", ""))
        rules.append(rec)

    by_class = {}
    for r in rules:
        by_class[r["impl_class"]] = by_class.get(r["impl_class"], 0) + 1

    # 每模块系数之和（权重和），从规则实测计算（= 扣分公式分母）
    weight_sums: dict[str, float] = {}
    for r in rules:
        mk = r["module_key"]
        weight_sums[mk] = round(weight_sums.get(mk, 0.0) + r["coefficient"], 4)

    return {
        "rule_version": rule_version,
        "source": str(Path(xlsx_path).name),
        "source_sheet": "立项门禁",
        "applies_to": "开发类 EQQR8082 A2.1",
        "severity_map": SEVERITY_MAP,
        "severity_coeff": SEVERITY_COEFF,
        "scoring": {
            "source": "开发项目立项门禁扣分权重标准_82条规则.xlsx（收口-1，2026-07-03）",
            "total": 100,
            "veto_severity": "阻断",
            "tier_thresholds": TIER_THRESHOLDS,
            "module_base_scores": MODULE_BASE_SCORES,
            "module_weight_sums": weight_sums,
        },
        "severity_overrides_from_weight_table": overrides,
        "count": len(rules),
        "count_by_class": by_class,
        "count_blocking": sum(1 for r in rules if r["blocking"]),
        "rules": rules,
    }


def _norm(v) -> str:
    if v is None:
        return ""
    return str(v).replace("　", " ").strip()


def main():
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    version = sys.argv[2] if len(sys.argv) > 2 else "2026-07-03"
    weight = Path(sys.argv[3]) if len(sys.argv) > 3 else DEFAULT_WEIGHT_XLSX
    reg = build(xlsx, version, weight)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(reg, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"wrote {OUT}")
    print(f"  rule_version={reg['rule_version']}  count={reg['count']}  blocking={reg['count_blocking']}")
    print(f"  by_class={reg['count_by_class']}  (A=可机器核 B=半自动 C=转人工)")
    ov = reg["severity_overrides_from_weight_table"]
    print(f"  权重表严重度覆盖 {len(ov)} 条: {ov}")


if __name__ == "__main__":
    main()
