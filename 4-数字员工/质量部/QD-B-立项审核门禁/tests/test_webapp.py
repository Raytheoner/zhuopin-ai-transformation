"""任务 9.1 验收：QD-B 最小 Web 服务（上传 xlsx → 报告页）。"""
from __future__ import annotations

from io import BytesIO

import pytest

from qd_b_gate.webapp import _secure_filename, create_app


@pytest.fixture()
def client(tmp_path):
    app = create_app(upload_dir=tmp_path / "uploads", audit_path=tmp_path / "audit.jsonl")
    app.config["TESTING"] = True
    return app.test_client()


def test_ping(client):
    r = client.get("/api/ping")
    assert r.status_code == 200
    assert r.get_json()["status"] == "ok"


def test_index_shows_disclaimer_and_trial_badge(client):
    r = client.get("/")
    assert r.status_code == 200
    body = r.get_data(as_text=True)
    assert "试用版" in body
    assert "评审委员会" in body
    assert "<form" in body


class TestSecureFilename:
    """队列 #108②：上传文件名过滤——保留中文可读性，滤路径穿越/控制字符/Windows 非法字符。"""

    def test_preserves_chinese_filename(self):
        assert _secure_filename("华丰立项申请书.xlsx") == "华丰立项申请书.xlsx"

    def test_strips_path_traversal_unix_style(self):
        assert _secure_filename("../../etc/passwd.xlsx") == "passwd.xlsx"

    def test_strips_path_traversal_windows_style(self):
        assert _secure_filename("..\\..\\Windows\\evil.xlsx") == "evil.xlsx"

    def test_strips_absolute_windows_path(self):
        assert _secure_filename("C:\\secret\\立项书.xlsx") == "立项书.xlsx"

    def test_replaces_windows_illegal_characters(self):
        result = _secure_filename('a<b>c:d"e|f?g*h.xlsx')
        assert not any(ch in result for ch in '<>:"|?*')

    def test_strips_control_characters(self):
        result = _secure_filename("立项书\x00\x01.xlsx")
        assert "\x00" not in result and "\x01" not in result

    def test_strips_trailing_dots_and_spaces(self):
        result = _secure_filename("立项书.xlsx   ...")
        assert not result.endswith(" ") and not result.endswith(".")

    def test_empty_after_filtering_falls_back_to_default(self):
        assert _secure_filename("////") == "upload"


def test_evaluate_with_malicious_filename_stays_inside_upload_dir(client, tmp_path):
    """恶意文件名（路径穿越+特殊字符）不应逃逸 upload_dir，也不应导致落盘失败。"""
    data = {"proposal": (BytesIO(b"not a real xlsx"), "../../../evil<>:.xlsx")}
    r = client.post("/evaluate", data=data, content_type="multipart/form-data")
    # 内容本身不是合法 xlsx，评估会失败（500），但落盘阶段不应抛异常/逃逸目录——
    # 断言 uploads 目录下只新增了一个受控命名的文件，且不在目录之外。
    assert r.status_code in (400, 500)
    uploads_dir = tmp_path / "uploads"
    saved = list(uploads_dir.glob("*evil*"))
    assert len(saved) == 1
    assert saved[0].parent == uploads_dir


def test_evaluate_rejects_missing_file(client):
    r = client.post("/evaluate", data={}, content_type="multipart/form-data")
    assert r.status_code == 400
    assert "请选择" in r.get_data(as_text=True)


def test_evaluate_rejects_non_xlsx(client):
    data = {"proposal": (BytesIO(b"not an excel file"), "proposal.txt")}
    r = client.post("/evaluate", data=data, content_type="multipart/form-data")
    assert r.status_code == 400
    assert "仅支持 .xlsx" in r.get_data(as_text=True)


def test_evaluate_surfaces_exception_instead_of_blank_500(client):
    """损坏的 xlsx（非真实 Excel 格式）应如实回显错误，而非空白 500。"""
    data = {"proposal": (BytesIO(b"garbage-not-a-real-xlsx"), "broken.xlsx")}
    r = client.post("/evaluate", data=data, content_type="multipart/form-data")
    assert r.status_code == 500
    assert "评估失败" in r.get_data(as_text=True)


class TestEvaluateWithHuafeng:
    """真实黄金样本端到端：上传→报告页，判定与黄金基准一致（不需要 mock）。"""

    def test_huafeng_upload_produces_expected_verdict(self, client, huafeng_path):
        with open(huafeng_path, "rb") as fh:
            data = {"proposal": (fh, "华丰天然气发动机EPA认证服务咨询项目立项申请书.xlsx")}
            r = client.post("/evaluate", data=data, content_type="multipart/form-data")
        assert r.status_code == 200
        body = r.get_data(as_text=True)
        assert "不合格" in body  # 华丰黄金基准真值 = 不合格
        assert "AI 预审建议" in body
        # 跨模块段逐条呈现（变更包 qd-b-cross-module-check）：既不空着不说明，
        # 也不再用一句"尚未实现"把已核过的 C01/C08/C09/C10 一并说成没核。
        assert "跨模块校验尚未实现" not in body
        assert "跨模块校验结果（C01–C10）" in body
        assert "C05" in body and "所属里程碑" in body  # 未实现项写具体原因
        assert "转人工待办项" in body

    def test_report_page_陈忱反馈重排要素齐全(self, client, huafeng_path):
        """陈忱#116 反馈三要素：13模块得分率/扣分明细/全量明细双筛选+仅看问题项/下载按钮。"""
        with open(huafeng_path, "rb") as fh:
            data = {"proposal": (fh, "华丰天然气发动机EPA认证服务咨询项目立项申请书.xlsx")}
            r = client.post("/evaluate", data=data, content_type="multipart/form-data")
        body = r.get_data(as_text=True)
        assert "各模块得分率一览（13 模块）" in body
        assert "扣分明细" in body
        assert "全量评审明细表（82 项）" in body
        assert "仅看问题项" in body
        assert "下载 Excel 评分表" in body
        assert 'id="flt-module"' in body and 'id="flt-status"' in body
        assert "不合格扣分" in body and "待改进扣分" in body

    def test_download_link_serves_valid_workbook(self, client, huafeng_path):
        import io
        import re

        import openpyxl

        with open(huafeng_path, "rb") as fh:
            data = {"proposal": (fh, "华丰天然气发动机EPA认证服务咨询项目立项申请书.xlsx")}
            r = client.post("/evaluate", data=data, content_type="multipart/form-data")
        body = r.get_data(as_text=True)
        m = re.search(r'href="(/download/[^"]+)"', body)
        assert m, "报告页应含下载链接"
        dl = client.get(m.group(1))
        assert dl.status_code == 200
        assert "spreadsheetml" in dl.headers["Content-Type"]
        wb = openpyxl.load_workbook(io.BytesIO(dl.data))
        assert wb.sheetnames == ["评审汇总", "评审明细", "扣分明细", "跨模块校验"]

    def test_download_route_rejects_path_traversal(self, client):
        r = client.get("/download/..%2F..%2Fwebapp.py")
        assert r.status_code == 404


class TestCrossModuleThreeCarriersAgree:
    """④跨模块段的三处呈现载体（文本报告/网页页/Excel）必须同源同结论。

    此前三处各自写死"C01–C10 未实现"文案，任一处改了另两处不会跟着变——
    本类把"三载体一致"钉成回归，防再次漂移（变更包 qd-b-cross-module-check）。
    """

    def _artifacts(self, client, huafeng_path, tmp_path):
        import io
        import re

        import openpyxl

        from qd_b_gate.evaluate import evaluate

        with open(huafeng_path, "rb") as fh:
            data = {"proposal": (fh, "华丰.xlsx")}
            r = client.post("/evaluate", data=data, content_type="multipart/form-data")
        body = r.get_data(as_text=True)
        dl = client.get(re.search(r'href="(/download/[^"]+)"', body).group(1))
        ws = openpyxl.load_workbook(io.BytesIO(dl.data))["跨模块校验"]
        xlsx_rows = {ws.cell(row=i, column=1).value: ws.cell(row=i, column=5).value
                     for i in range(4, 14)}
        result = evaluate(huafeng_path, audit_path=tmp_path / "a.jsonl", sample_id="华丰")
        return body, xlsx_rows, result.report

    def test_same_ten_checks_and_same_verdicts_across_carriers(
            self, client, huafeng_path, tmp_path):
        from qd_b_gate.report_items import STATUS_LABELS

        body, xlsx_rows, report = self._artifacts(client, huafeng_path, tmp_path)
        expected = {i.rule_id: STATUS_LABELS[i.verdict] for i in report.cross_module_items}
        assert xlsx_rows == expected
        text = report.to_text()
        for check_id, label in expected.items():
            assert check_id in body and check_id in text
            assert f"[{label}]" in text
