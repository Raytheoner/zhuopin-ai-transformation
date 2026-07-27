"""评审报告 Excel 导出（陈忱灰度反馈 #116 ③）。

3 sheet：评审汇总 / 评审明细 / 扣分明细，格式基准=陈忱回件模板
`7-外部文档/质量部/质量部-ChenChen-回复-2026-07-24-QD-B立项审核门禁_评审报告模板-*.xlsx`
（sheet 名/列头/合并单元格布局对齐该模板；状态颜色/权重口径见下）。

仅展示层：逐项标准分/实得分/扣分复用 report_items.py 的口径（与 scoring.py 判据
完全一致），**不改 80 条权重表/扣分判据**（口径红线，须走判例批改+陈忱显式签认）。
"""
from __future__ import annotations

import re
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .evaluate import EvaluationResult
from .models import Verdict
from .report import GateReport
from .report_items import (
    ScoredItem,
    build_basic_info,
    build_financial_summary,
    build_scored_items,
    deduction_items,
    deduction_subtotals,
)
from .rules.registry import RuleRegistry, load_registry

_HEADER_FILL = PatternFill("solid", fgColor="2C3E50")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_MODULE_FILL = PatternFill("solid", fgColor="D9E1F2")
_MODULE_FONT = Font(bold=True)
_TITLE_FONT = Font(bold=True, size=14)
_LABEL_FONT = Font(bold=True)
_THIN = Side(style="thin", color="B0B0B0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# 状态颜色标识（陈忱反馈原话：通过绿/待改进黄/不合格红/不适用灰；
# 转人工/未实现是本系统如实标注的额外两态，不在陈忱原始四态内，给中性配色区分）
_STATUS_COLORS: dict[Verdict, tuple[str, str]] = {
    Verdict.PASS: ("C6EFCE", "006100"),
    Verdict.NA: ("D9D9D9", "595959"),
    Verdict.WARN: ("FFEB9C", "9C6500"),
    Verdict.FAIL: ("FFC7CE", "9C0006"),
    Verdict.MANUAL: ("DDEBF7", "1F4E78"),
    Verdict.PENDING: ("F2F2F2", "808080"),
}

_DETAIL_COLUMNS = ["序号", "所属模块", "检查项", "评审标准", "状态", "权重", "标准分", "实际得分", "扣分", "详情"]
_DETAIL_WIDTHS = [6, 20, 32, 34, 10, 8, 8, 10, 8, 46]

_FORBIDDEN_FS_CHARS = re.compile(r'[\\/:*?"<>|\r\n\t]')


def _sanitize_filename_part(text: str) -> str:
    text = (text or "").strip()
    text = _FORBIDDEN_FS_CHARS.sub("_", text)
    text = re.sub(r"\s+", "", text)
    return text[:60] or "未命名项目"


def report_filename(result: EvaluationResult, when: datetime | None = None) -> str:
    """`项目名称_评审报告_YYYYMMDD.xlsx`（陈忱反馈原文件名格式）。"""
    doc = result.document
    fv = doc.get("一、项目信息/项目名称")
    project = fv.value if fv.is_present else (result.report.sample_id or "未命名项目")
    when = when or datetime.now()
    return f"{_sanitize_filename_part(str(project))}_评审报告_{when:%Y%m%d}.xlsx"


def _set_col_widths(ws: Worksheet, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _title_row(ws: Worksheet, text: str, ncols: int) -> None:
    ws.cell(row=1, column=1, value=text)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1)
    c.font = _TITLE_FONT
    c.alignment = Alignment(horizontal="center")


def _write_header_row(ws: Worksheet, row_idx: int, columns: list[str]) -> None:
    for col_idx, text in enumerate(columns, start=1):
        c = ws.cell(row=row_idx, column=col_idx, value=text)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _BORDER


def _write_module_divider(ws: Worksheet, row_idx: int, text: str, ncols: int) -> None:
    ws.cell(row=row_idx, column=1, value=text)
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=ncols)
    c = ws.cell(row=row_idx, column=1)
    c.fill = _MODULE_FILL
    c.font = _MODULE_FONT


def _write_item_row(ws: Worksheet, row_idx: int, item: ScoredItem) -> None:
    values = [item.idx, item.section, item.check_item, item.pass_condition,
              item.status_label, item.coefficient, item.std_score,
              item.actual_score, item.deduction, item.detail_text]
    fill_color, font_color = _STATUS_COLORS.get(item.verdict, ("FFFFFF", "000000"))
    for col_idx, v in enumerate(values, start=1):
        c = ws.cell(row=row_idx, column=col_idx, value=v)
        c.border = _BORDER
        c.alignment = Alignment(vertical="top", wrap_text=col_idx in (3, 4, 10))
        if col_idx == 5:
            c.fill = PatternFill("solid", fgColor=fill_color)
            c.font = Font(color=font_color, bold=True)


def _build_summary_sheet(wb: Workbook, result: EvaluationResult, registry: RuleRegistry,
                         items: list[ScoredItem], when: datetime) -> None:
    ws = wb.create_sheet("评审汇总")
    _set_col_widths(ws, [16, 16, 16, 16, 16])
    report = result.report
    sr = report.score_result
    fail_ded, warn_ded = deduction_subtotals(items)

    ws.cell(row=1, column=1, value="立项申请书合格性评审报告")
    ws.merge_cells("A1:E1")
    ws["A1"].font = _TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.cell(row=2, column=1, value=f"文件: {report.sample_id or '(未命名)'}")
    ws.cell(row=2, column=2, value=f"评审时间: {when:%Y-%m-%d %H:%M}")

    score_text = "—（一票否决）" if sr.veto else f"{sr.total_score:.1f}"
    ws.cell(row=4, column=1, value=f"判定结果: {report.verdict}    得分: {score_text} / 100")
    ws["A4"].font = _LABEL_FONT

    _write_header_row(ws, 6, ["满分", "实际得分", "不合格扣分", "待改进扣分"])
    ws.cell(row=7, column=1, value="100分")
    ws.cell(row=7, column=2, value="—" if sr.veto else f"{sr.total_score:.1f}分")
    ws.cell(row=7, column=3, value=f"-{fail_ded:.1f}分")
    ws.cell(row=7, column=4, value=f"-{warn_ded:.1f}分")
    for col in range(1, 5):
        ws.cell(row=7, column=col).border = _BORDER

    if sr.provisional:
        ws.cell(row=8, column=1,
                value=f"⚠ 暂定：{sr.pending} 条 A 类规则未实现（视为通过），全量实现后复核")

    ws.cell(row=9, column=1, value="项目基本信息")
    ws.merge_cells("A9:E9")
    ws["A9"].font = _LABEL_FONT
    ws["A9"].fill = _MODULE_FILL

    info = build_basic_info(result.document)
    ws.cell(row=10, column=1, value="项目名称").font = _LABEL_FONT
    ws.merge_cells("B10:C10")
    ws.cell(row=10, column=2, value=info["项目名称"])
    ws.cell(row=10, column=4, value="客户名称").font = _LABEL_FONT
    ws.cell(row=10, column=5, value=info["客户名称"])

    ws.cell(row=11, column=1, value="项目经理").font = _LABEL_FONT
    ws.merge_cells("B11:C11")
    ws.cell(row=11, column=2, value=info["项目经理"])
    ws.cell(row=11, column=4, value="项目类型").font = _LABEL_FONT
    ws.cell(row=11, column=5, value=info["项目类型"])

    ws.cell(row=12, column=1, value="所属事业部").font = _LABEL_FONT
    ws.merge_cells("B12:C12")
    ws.cell(row=12, column=2, value=info["所属事业部"])
    ws.cell(row=12, column=4, value="项目编号").font = _LABEL_FONT
    ws.cell(row=12, column=5, value=info["项目编号"])

    ws.cell(row=14, column=1, value="财务摘要")
    ws.merge_cells("A14:E14")
    ws["A14"].font = _LABEL_FONT
    ws["A14"].fill = _MODULE_FILL

    fin = build_financial_summary(result.document)
    ws.cell(row=15, column=1, value="项目收入").font = _LABEL_FONT
    ws.cell(row=15, column=2, value="项目成本").font = _LABEL_FONT
    ws.cell(row=15, column=3, value="毛利率").font = _LABEL_FONT
    ws.cell(row=15, column=4, value="收益指标").font = _LABEL_FONT
    ws.merge_cells("D15:E15")
    ws.cell(row=16, column=1, value=fin["项目收入"])
    ws.cell(row=16, column=2, value=fin["项目成本"])
    ws.cell(row=16, column=3, value=fin["毛利率"])
    ws.merge_cells("D16:E16")
    ws.cell(row=16, column=4, value=fin["收益指标"])

    ws.cell(row=18, column=1, value=report.disclaimer)
    ws["A18"].font = Font(italic=True, color="9C6500")
    ws.merge_cells("A18:E18")


def _build_detail_sheet(wb: Workbook, items: list[ScoredItem],
                        report: GateReport, ncols: int = 10) -> None:
    ws = wb.create_sheet("评审明细")
    _set_col_widths(ws, _DETAIL_WIDTHS)
    _title_row(ws, "立项申请书合格性评审明细", ncols)

    header_row = 3
    _write_header_row(ws, header_row, _DETAIL_COLUMNS)

    module_scores = report.score_result.module_scores
    row_idx = header_row + 1
    current_module = None
    for item in items:
        if item.module_key != current_module:
            current_module = item.module_key
            ms = module_scores.get(current_module)
            rate_text = f"（得分率 {ms.score / ms.base * 100:.1f}%）" if ms and ms.base else ""
            _write_module_divider(ws, row_idx, f"{item.section} {rate_text}", ncols)
            row_idx += 1
        _write_item_row(ws, row_idx, item)
        row_idx += 1

    last_row = row_idx - 1
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ncols)}{last_row}"


def _build_deduction_sheet(wb: Workbook, items: list[ScoredItem], ncols: int = 10) -> None:
    ws = wb.create_sheet("扣分明细")
    _set_col_widths(ws, _DETAIL_WIDTHS)
    _title_row(ws, "扣分明细（仅展示待改进/不合格项，按扣分从高到低排序）", ncols)

    header_row = 3
    _write_header_row(ws, header_row, _DETAIL_COLUMNS)

    rows = deduction_items(items)
    if not rows:
        ws.cell(row=4, column=1, value="无待改进/不合格项——本次评审零扣分")
        ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=ncols)
        ws.freeze_panes = f"A{header_row + 1}"
        return

    row_idx = header_row + 1
    for item in rows:
        _write_item_row(ws, row_idx, item)
        row_idx += 1

    last_row = row_idx - 1
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ncols)}{last_row}"


def build_workbook(result: EvaluationResult, registry: RuleRegistry | None = None,
                   when: datetime | None = None) -> Workbook:
    """构建三 sheet 评审报告工作簿（评审汇总/评审明细/扣分明细）。"""
    reg = registry or load_registry()
    items = build_scored_items(result.report, reg)
    when = when or datetime.now()

    wb = Workbook()
    wb.remove(wb.active)  # 默认空白 sheet，用具名 sheet 替代
    _build_summary_sheet(wb, result, reg, items, when)
    _build_detail_sheet(wb, items, result.report)
    _build_deduction_sheet(wb, items)
    return wb
