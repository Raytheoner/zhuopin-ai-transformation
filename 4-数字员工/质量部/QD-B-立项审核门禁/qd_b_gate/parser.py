"""立项申请书解析器（intake-parser）。

把 EQQR8082 A2.1 合并单元格 Excel 表单解析为 ProposalDocument。

设计纪律（design.md D3/D4）：
- 取数锚「章节标题文本」，不锚绝对单元格坐标（模板一月四版）。
- 合并单元格按左上锚点值展开。
- 字段抽取状态显式标记，区分"业务空"与"解析未命中"。

doc_parser 当前仅 QD-B 一个真实消费方（QD-A/8D 尚未建），按 rule-of-three
暂留本场景，预留干净接口；第 2 真实消费方出现再提升进 shared_tools。
"""
from __future__ import annotations

import re
from pathlib import Path

import openpyxl
from openpyxl.utils import coordinate_to_tuple, get_column_letter
from openpyxl.utils.datetime import from_excel

from .models import ExtractStatus, FieldValue, ProposalDocument

# 13 模块「序号 + 规范名」。**取数锚序号前缀（一、…十三、），不锚描述文本**——
# 描述文本随模板版本变（如模块八："项目成本及收益分析" / "项目全生命周期的成本效益
# 分析"；模块九："项目里程碑计划与阶段预算规划" / "项目开发期限、时间节点及各阶段
# 预算"），但序号前缀稳定（design.md D3 抗改版）。规范名仅用于字段键/展示。
SECTION_ORDINALS = ["一", "二", "三", "四", "五", "六", "七", "八", "九",
                    "十", "十一", "十二", "十三"]
SECTION_TITLES = [
    "一、项目信息",
    "二、立项依据",
    "三、项目的目的和意义",
    "四、项目目标",
    "五、关键技术与开发能力评估",
    "六、项目风险分析",
    "七、项目所需资源采购计划",
    "八、项目成本及收益分析",
    "九、项目里程碑计划与阶段预算规划",
    "十、项目现金流分析",
    "十一、项目团队成员",
    "十二、总结",
    "十三、立项决议",
]
# 序号 -> 规范标题键（按位置对应）
_ORD_TO_TITLE = dict(zip(SECTION_ORDINALS, SECTION_TITLES))


def _norm(s) -> str:
    """规范化文本：去空白、全角空格、首尾标点空格，便于锚点匹配。"""
    if s is None:
        return ""
    return str(s).replace("　", " ").strip()


def _num(v):
    """从单元格值提取数值（人月/金额等）。取首个数字，失败返回 None。"""
    if isinstance(v, (int, float)):
        return float(v)
    m = re.search(r"-?\d+(?:\.\d+)?", _norm(v))
    return float(m.group()) if m else None


class ProposalParser:
    """按章节标题锚点解析立项书。"""

    def __init__(self, path: str | Path):
        self.path = str(path)
        self.wb = openpyxl.load_workbook(path, data_only=True)
        self.ws = self._pick_sheet()
        # 合并区域：左上坐标 -> range 对象；坐标 -> 左上坐标（值解析用）
        self._merge_anchor: dict[str, str] = {}
        for rng in self.ws.merged_cells.ranges:
            tl = rng.coord.split(":")[0]
            for row in range(rng.min_row, rng.max_row + 1):
                for col in range(rng.min_col, rng.max_col + 1):
                    self._merge_anchor[f"{get_column_letter(col)}{row}"] = tl

    def _pick_sheet(self):
        for name in self.wb.sheetnames:
            if "立项申请书" in name and "开发" in name:
                return self.wb[name]
        # 退路：最大的工作表
        return max(self.wb.worksheets, key=lambda w: (w.max_row or 0) * (w.max_column or 0))

    # ---------- 基础取值 ----------
    def cell(self, coord: str):
        """读单元格，自动解析合并区域到左上锚点值。"""
        anchor = self._merge_anchor.get(coord, coord)
        return self.ws[anchor].value

    def _value_adjacent(self, coord: str) -> tuple[object, str]:
        """读 label 单元格紧邻右侧的「值单元格」（label 合并跨度结束列+1）。

        关键纪律（解析探针经验）：取「紧邻」单元格而非「向右第一个非空」——
        否则空白模板里值为空时会误抓到右边的下一个 label。值为空 → 返回
        (None, 坐标)，由调用方判为 MISSING（业务空），与 NOT_FOUND 区分。
        """
        r, c = coordinate_to_tuple(coord)
        start = c + 1
        rng = self._range_of(coord)
        if rng is not None:
            start = rng.max_col + 1
        cc = f"{get_column_letter(start)}{r}"
        v = self.cell(cc)
        src = self._merge_anchor.get(cc, cc)
        if v is None or _norm(v) == "":
            return None, src
        return v, src

    def _range_of(self, coord: str):
        anchor = self._merge_anchor.get(coord, coord)
        for rng in self.ws.merged_cells.ranges:
            if rng.coord.split(":")[0] == anchor:
                return rng
        return None

    def find_label(self, text: str, row_range: tuple[int, int] | None = None) -> str | None:
        """在（可选行范围内）按文本找 label 单元格坐标。以 norm 后「相等或以之开头」匹配。"""
        t = _norm(text)
        lo, hi = row_range or (1, self.ws.max_row)
        for row in self.ws.iter_rows(min_row=lo, max_row=hi):
            for cell in row:
                cv = _norm(cell.value)
                if cv and (cv == t or cv.startswith(t)):
                    return cell.coordinate
        return None

    # ---------- 章节定位 ----------
    def section_rows(self) -> dict[str, tuple[int, int]]:
        """返回各章节（规范标题）-> (起始行, 结束行)。

        **按 A 列序号前缀（一、…十三、）定位**，吸收描述文本的版本差异（D3）。
        结束行 = 下一章节起始行-1。
        """
        # 序号 -> 行号（取该序号在 A 列首次出现）
        ord_row: dict[str, int] = {}
        for row in self.ws.iter_rows():
            cell = row[0]            # A 列
            t = _norm(cell.value)
            if not t:
                continue
            m = re.match(r"^(十[一二三]?|[一二三四五六七八九])[、，]", t)
            if m and m.group(1) in _ORD_TO_TITLE and m.group(1) not in ord_row:
                ord_row[m.group(1)] = cell.row
        starts = sorted(((_ORD_TO_TITLE[o], r) for o, r in ord_row.items()),
                        key=lambda x: x[1])
        out: dict[str, tuple[int, int]] = {}
        for i, (title, row) in enumerate(starts):
            end = starts[i + 1][1] - 1 if i + 1 < len(starts) else self.ws.max_row
            out[title] = (row, end)
        return out

    def detect_template_version(self) -> str:
        """识别模板版本（A2.1 / A2 / A1 / A0）。

        来源优先级：封面/履历单元格 → 文件名。吸收"A2.1""2.1版本""测试专用2.1"
        等写法（华丰填件文件名即"…2.1版本"，无"A"前缀）。
        """
        def classify(text: str) -> str | None:
            t = _norm(text)
            if not t:
                return None
            if re.search(r"A?2\.1|2\.1\s*版本", t):
                return "A2.1"
            if re.search(r"\bA2\b|A2版", t):
                return "A2"
            if re.search(r"\bA1\b", t):
                return "A1"
            if re.search(r"\bA0\b", t):
                return "A0"
            return None

        # 1) 优先「当前版本」显式标记（封面），最权威
        for name in self.wb.sheetnames:
            ws = self.wb[name]
            coord = self.find_label("当前版本") if ws is self.ws else None
            for row in ws.iter_rows(min_row=1, max_row=min(40, ws.max_row or 1)):
                for cell in row:
                    if _norm(cell.value) == "当前版本":
                        # 取右侧/下方相邻值
                        for nb in (cell.row, cell.row + 1):
                            for nc in range(cell.column, cell.column + 4):
                                vv = classify(ws.cell(row=nb, column=nc).value)
                                if vv:
                                    return vv
        # 2) 退路：扫封面/履历任意单元格
        for name in self.wb.sheetnames:
            ws = self.wb[name]
            for row in ws.iter_rows(min_row=1, max_row=min(40, ws.max_row or 1)):
                for cell in row:
                    v = classify(cell.value)
                    if v:
                        return v
        # 3) 末路：文件名
        return classify(Path(self.path).name) or "unknown"

    def filename_version(self) -> str:
        """文件名暗示的版本（与封面嵌入版本比对，揭示命名/版本不一致）。"""
        t = _norm(Path(self.path).name)
        if re.search(r"2\.1", t):
            return "A2.1"
        if re.search(r"\bA2\b|A2版", t):
            return "A2"
        return "unknown"

    @staticmethod
    def _maybe_date(value):
        """Excel 日期序列号 → YYYY-MM-DD。非日期序列原样返回。

        填件里日期常以序列号缓存（如 46188）。合理区间 [25569(1970), 80000(~2119)]
        视为日期序列，转 ISO 日期串；否则原样。
        """
        if isinstance(value, (int, float)) and 25569 <= value <= 80000:
            try:
                return from_excel(value).date().isoformat()
            except Exception:
                return value
        return value

    # ---------- 主入口 ----------
    def parse(self) -> ProposalDocument:
        doc = ProposalDocument(source_path=self.path)
        doc.template_version = self.detect_template_version()
        fname_ver = self.filename_version()
        if doc.template_version != "A2.1":
            doc.warnings.append(
                f"封面嵌入版本={doc.template_version}（MVP 目标 A2.1）；序号锚点解析与版本无关、"
                f"已正常定位，但版本差异涉及的模块（如采购计划）字段需抽样核对"
            )
        if fname_ver != "unknown" and fname_ver != doc.template_version:
            doc.warnings.append(
                f"⚠ 版本不一致：文件名暗示 {fname_ver}，封面「当前版本」为 {doc.template_version}"
                f"——交陈忱确认该填件实际版本（命名/封面/履历对齐）"
            )
        secs = self.section_rows()
        missing_secs = [t for t in SECTION_TITLES if t not in secs]
        if missing_secs:
            doc.warnings.append(f"未定位到章节：{missing_secs}")

        self._parse_project_info(doc, secs.get("一、项目信息"))
        # 项目类型驱动适用规则集
        pt = doc.fields.get("一、项目信息/项目类型")
        if pt and pt.is_present:
            doc.project_type = str(pt.value).strip()
        self._parse_personnel_table(doc, secs.get("一、项目信息"))
        self._parse_text_areas(doc, secs)
        self._parse_risk_table(doc, secs.get("六、项目风险分析"))
        self._parse_resource_table(doc, secs.get("七、项目所需资源采购计划"))
        self._parse_cost_benefit_section(doc, secs.get("八、项目成本及收益分析"))
        self._parse_budget_table(doc, secs.get("九、项目里程碑计划与阶段预算规划"))
        self._parse_cashflow_table(doc, secs.get("十、项目现金流分析"))
        return doc

    # ---------- 表头行按文本定位列（财务表通用） ----------
    def _row_label_cols(self, row: int, labels: list[str]) -> dict[str, int]:
        """在指定行内按文本前缀匹配定位各列列号（同一行内表头文本互不重叠）。"""
        out: dict[str, int] = {}
        if row < 1:
            return out
        for col in range(1, self.ws.max_column + 1):
            t = _norm(self.ws.cell(row=row, column=col).value)
            if not t:
                continue
            for lab in labels:
                if lab not in out and t.startswith(lab):
                    out[lab] = col
        return out

    def _cell_num(self, row: int, col: int | None):
        if not col:
            return None
        return _num(self.ws.cell(row=row, column=col).value)

    def _cell_text(self, row: int, col: int | None) -> str:
        if not col:
            return ""
        return _norm(self.ws.cell(row=row, column=col).value)

    # ---------- 模块六：项目风险分析 ----------
    def _parse_risk_table(self, doc: ProposalDocument, rng: tuple[int, int] | None):
        if rng is None:
            return
        lo, hi = rng
        coord = self.find_label("风险系数", (lo, hi))
        if not coord:
            doc.warnings.append("模块六风险表头（风险系数列）未命中")
            return
        hr = coordinate_to_tuple(coord)[0]
        cols = self._row_label_cols(hr, ["类别", "描述", "严重度", "发生频率",
                                         "风险系数", "风险等级", "应对措施"])
        rows: list[dict] = []
        for r in range(hr + 1, hi + 1):
            if self.ws.cell(row=r, column=1).value in (None, ""):
                continue
            rows.append({
                "序号": self.ws.cell(row=r, column=1).value,
                "类别": self._cell_text(r, cols.get("类别")),
                "描述": self._cell_text(r, cols.get("描述")),
                "严重度": self._cell_text(r, cols.get("严重度")),
                "频率": self._cell_text(r, cols.get("发生频率")),
                "系数": self._cell_num(r, cols.get("风险系数")),
                "等级": self._cell_text(r, cols.get("风险等级")),
                "应对措施": self._cell_text(r, cols.get("应对措施")),
            })
        doc.tables["风险"] = rows

    # ---------- 模块七：项目所需资源采购计划 ----------
    def _parse_resource_table(self, doc: ProposalDocument, rng: tuple[int, int] | None):
        if rng is None:
            return
        lo, hi = rng
        coord = self.find_label("获取方式", (lo, hi))
        if not coord:
            doc.warnings.append("模块七资源表头（获取方式列）未命中")
            return
        hr = coordinate_to_tuple(coord)[0]
        cols = self._row_label_cols(hr, ["类型", "原因", "主要配置", "所需数量",
                                         "获取方式", "推荐供应商", "预算"])
        rows: list[dict] = []
        for r in range(hr + 1, hi + 1):
            t = self._cell_text(r, cols.get("类型", 1))
            if not t:
                continue
            rows.append({
                "类型": t,
                "原因": self._cell_text(r, cols.get("原因")),
                "配置": self._cell_text(r, cols.get("主要配置")),
                "数量": self._cell_num(r, cols.get("所需数量")),
                "获取方式": self._cell_text(r, cols.get("获取方式")),
                "供应商": self._cell_text(r, cols.get("推荐供应商")),
                "预算": self._cell_num(r, cols.get("预算")),
                "预算文本": self._cell_text(r, cols.get("预算")),
            })
        doc.tables["资源"] = rows

    # ---------- 模块八：成本及收益分析（(一)成本效益表 + (二)收益分析说明） ----------
    _CB_COLS = ["收入D", "成本E", "利润 F=D-E", "数量G", "销售单价H", "成本单价I",
                "收入J=G*H", "成本K=G*I", "三包质量费L", "利润 M=J-K-L"]
    _CB_KEYMAP = {  # 表头前缀 -> 规范字段键
        "收入D": "收入D", "成本E": "成本E", "利润 F=D-E": "利润F",
        "数量G": "数量G", "销售单价H": "单价H", "成本单价I": "成本单价I",
        "收入J=G*H": "收入J", "成本K=G*I": "成本K",
        "三包质量费L": "三包L", "利润 M=J-K-L": "利润M",
    }
    _INCOME_SCALARS = {
        "产品量纲预估依据": "量纲预估依据",
        "产品售价预估依据": "售价预估依据",
        "产品量产成本预估依据": "量产成本预估依据",
        "项目本身毛利率": "项目本身毛利率",
        "项目全生命周期毛利率": "全生命周期毛利率",
        "技术服务类（项目收入": "收益指标_技术服务类",
        "产品类（（项目收入": "收益指标_产品类",
    }

    def _parse_cost_benefit_section(self, doc: ProposalDocument, rng: tuple[int, int] | None):
        if rng is None:
            return
        lo, hi = rng
        coord2 = self.find_label("（二）", (lo, hi))
        end_cb = (coordinate_to_tuple(coord2)[0] - 1) if coord2 else hi

        self._parse_cost_benefit_table(doc, lo, end_cb)
        if coord2:
            start_two = coordinate_to_tuple(coord2)[0]
            self._parse_revenue_basis_table(doc, start_two, hi)
            self._parse_income_analysis_scalars(doc, start_two, hi)

    def _parse_cost_benefit_table(self, doc: ProposalDocument, lo: int, hi: int):
        coord = self.find_label("数量G", (lo, hi))
        if not coord:
            doc.warnings.append("模块八（一）成本效益表头（数量G列）未命中")
            return
        hr = coordinate_to_tuple(coord)[0]
        cols = self._row_label_cols(hr, self._CB_COLS)
        p_cols = self._row_label_cols(hr - 1, ["利润小计P=F+M"])
        p_col = p_cols.get("利润小计P=F+M")

        rows: list[dict] = []
        for r in range(hr + 1, hi + 1):
            label = self._cell_text(r, 1)
            if not label:
                continue
            row: dict = {"年份": label, "is_total": label == "合计"}
            for hdr, key in self._CB_KEYMAP.items():
                row[key] = self._cell_num(r, cols.get(hdr))
            row["利润小计P"] = self._cell_num(r, p_col)
            rows.append(row)
        doc.tables["成本效益"] = rows

    def _parse_revenue_basis_table(self, doc: ProposalDocument, lo: int, hi: int):
        coord = self.find_label("对应商务阶段标志", (lo, hi))
        if not coord:
            doc.warnings.append("模块八（二）收入依据表头（商务阶段标志列）未命中")
            return
        hr = coordinate_to_tuple(coord)[0]
        cols = self._row_label_cols(hr, ["对应商务阶段标志", "合同约定/预计收款金额",
                                         "合同约定/预计收款时间"])
        rows: list[dict] = []
        total = None
        for r in range(hr + 1, hi + 1):
            label = self._cell_text(r, 1)
            if not label:
                continue
            if label == "合计":
                total = self._cell_num(r, cols.get("合同约定/预计收款金额"))
                break
            rows.append({
                "阶段": label,
                "商务标志": self._cell_text(r, cols.get("对应商务阶段标志")),
                "收款金额": self._cell_num(r, cols.get("合同约定/预计收款金额")),
                "收款时间": self._cell_text(r, cols.get("合同约定/预计收款时间")),
            })
        doc.tables["收入依据"] = rows
        if total is not None:
            doc.tables["收入依据合计"] = [{"合计": total}]

    def _parse_income_analysis_scalars(self, doc: ProposalDocument, lo: int, hi: int):
        for prefix, key in self._INCOME_SCALARS.items():
            coord = self.find_label(prefix, (lo, hi))
            fkey = f"八、（二）收益分析说明/{key}"
            if coord is None:
                doc.fields[fkey] = FieldValue(key=fkey, status=ExtractStatus.NOT_FOUND, anchor=prefix)
                continue
            value, src = self._value_adjacent(coord)
            if value is None:
                doc.fields[fkey] = FieldValue(key=fkey, status=ExtractStatus.MISSING,
                                              anchor=prefix, source_cell=src)
            else:
                doc.fields[fkey] = FieldValue(key=fkey, value=_norm(value),
                                              status=ExtractStatus.EXTRACTED,
                                              anchor=prefix, source_cell=src)

    # ---------- 模块九：开发期限及预算 ----------
    def _parse_budget_table(self, doc: ProposalDocument, rng: tuple[int, int] | None):
        if rng is None:
            return
        lo, hi = rng
        coord = self.find_label("小计⑩", (lo, hi))
        if not coord:
            doc.warnings.append("模块九预算表头（小计⑩列）未命中")
            return
        hr = coordinate_to_tuple(coord)[0]
        cols = self._row_label_cols(hr, ["人员预估成本②", "设备&模具③",
                                         "委外开发费&实验费&咨询费④", "材料费&打样费⑤",
                                         "差旅费&招待费⑥", "折旧摊销&福利费⑦", "其他⑧",
                                         "项目售后费用计提⑨", "小计⑩"])
        sub = self._row_label_cols(hr - 1, ["开始时间", "结束时间", "人月数"])
        top = self._row_label_cols(hr - 2, ["项目收入预算估计", "现金流情况"])

        rows: list[dict] = []
        for r in range(hr + 1, hi + 1):
            stage = self._cell_text(r, 1)
            if not stage:
                continue
            rows.append({
                "阶段": stage, "is_total": stage == "合计",
                "开始": self.ws.cell(row=r, column=sub["开始时间"]).value if sub.get("开始时间") else None,
                "结束": self.ws.cell(row=r, column=sub["结束时间"]).value if sub.get("结束时间") else None,
                "人月": self._cell_num(r, sub.get("人月数")),
                "人员成本②": self._cell_num(r, cols.get("人员预估成本②")),
                "设备③": self._cell_num(r, cols.get("设备&模具③")),
                "委外④": self._cell_num(r, cols.get("委外开发费&实验费&咨询费④")),
                "材料⑤": self._cell_num(r, cols.get("材料费&打样费⑤")),
                "差旅⑥": self._cell_num(r, cols.get("差旅费&招待费⑥")),
                "折旧⑦": self._cell_num(r, cols.get("折旧摊销&福利费⑦")),
                "其他⑧": self._cell_num(r, cols.get("其他⑧")),
                "售后⑨": self._cell_num(r, cols.get("项目售后费用计提⑨")),
                "小计⑩": self._cell_num(r, cols.get("小计⑩")),
                "收入⑪": self._cell_num(r, top.get("项目收入预算估计")),
                "现金流": self._cell_num(r, top.get("现金流情况")),
            })
        doc.tables["预算表"] = rows

    # ---------- 模块十：项目现金流分析 ----------
    def _parse_cashflow_table(self, doc: ProposalDocument, rng: tuple[int, int] | None):
        if rng is None:
            return
        lo, hi = rng
        coord = self.find_label("累计现金流", (lo, hi))
        if not coord:
            doc.warnings.append("模块十现金流表头（累计现金流列）未命中")
            return
        hr = coordinate_to_tuple(coord)[0]
        cols = self._row_label_cols(hr, ["收入", "支出-人力", "支出-设备", "支出-委外",
                                         "支出-材料", "支出-差旅", "支出-折旧摊销",
                                         "支出-其他", "支出-项目售后费用计提",
                                         "月净现金流", "累计现金流"])
        rows: list[dict] = []
        for r in range(hr + 1, hi + 1):
            month = self.ws.cell(row=r, column=1).value
            if month in (None, ""):
                continue
            rows.append({
                "月份": month,
                "收入": self._cell_num(r, cols.get("收入")),
                "人力": self._cell_num(r, cols.get("支出-人力")),
                "设备": self._cell_num(r, cols.get("支出-设备")),
                "委外": self._cell_num(r, cols.get("支出-委外")),
                "材料": self._cell_num(r, cols.get("支出-材料")),
                "差旅": self._cell_num(r, cols.get("支出-差旅")),
                "折旧摊销": self._cell_num(r, cols.get("支出-折旧摊销")),
                "其他": self._cell_num(r, cols.get("支出-其他")),
                "售后": self._cell_num(r, cols.get("支出-项目售后费用计提")),
                "月净现金流": self._cell_num(r, cols.get("月净现金流")),
                "累计现金流": self._cell_num(r, cols.get("累计现金流")),
            })
        doc.tables["现金流表"] = rows

    # ---------- 模块一：项目信息 ----------
    # label 文本 -> 规范字段键（吸收尾随空格）
    INFO_FIELDS = {
        "项目名称": "项目名称",
        "项目令": "项目令",
        "客户CTS/SOR版本号": "客户CTS/SOR版本号",
        "项目代号": "项目代号",
        "项目经理": "项目经理",
        "目标车型/车辆平台": "目标车型/车辆平台",
        "客户名称": "客户名称",
        "项目所属事业部": "项目所属事业部",
        "SOP目标年份": "SOP目标年份/项目结项时间",
        "开始日期": "开始日期",
        "结束日期": "结束日期",
        "功能安全目标ASIL": "功能安全目标ASIL",
        "预算总工时": "预算总工时",
        "项目类型": "项目类型",          # 产品类/技术服务类（文本/单选，非 True/False 复选框）
    }

    # 复选框字段（True/False 成对）。项目类型不在此列 —— 它是文本单选，走普通取值。
    CHECKBOX_FIELDS = ["项目等级", "适用法规/体系"]

    # 日期型字段：抽到 Excel 序列号时转 ISO 日期
    DATE_FIELDS = {"开始日期", "结束日期"}

    def _parse_project_info(self, doc: ProposalDocument, rng: tuple[int, int] | None):
        if rng is None:
            doc.warnings.append("模块一未定位，项目信息字段全部 NOT_FOUND")
            return
        for label_text, key in self.INFO_FIELDS.items():
            coord = self.find_label(label_text, rng)
            fkey = f"一、项目信息/{key}"
            if coord is None:
                doc.fields[fkey] = FieldValue(key=fkey, status=ExtractStatus.NOT_FOUND,
                                              anchor=label_text, reason="标题锚点未命中")
                continue
            value, src = self._value_adjacent(coord)
            if value is None:
                doc.fields[fkey] = FieldValue(key=fkey, status=ExtractStatus.MISSING,
                                              anchor=label_text, source_cell=src,
                                              reason="锚点命中但值为空")
            else:
                if key in self.DATE_FIELDS:
                    value = self._maybe_date(value)
                doc.fields[fkey] = FieldValue(key=fkey, value=_norm(value),
                                              status=ExtractStatus.EXTRACTED,
                                              anchor=label_text, source_cell=src)

        # 勾选项：项目等级 / 适用法规体系 — True/False 成对，扫描止于下一 label
        for cb in self.CHECKBOX_FIELDS:
            self._parse_checkbox_row(doc, cb, rng)

    # ---------- 文本区域模块（立项依据 / 目的意义 / 总结） ----------
    # 模块 -> 该模块「说明/提示」行文本前缀（用于跳过，避免把空白模板的填写指引当内容）
    _TEXT_MODULES = {
        "二、立项依据": ["简洁", "简明", "说明本项目"],
        "三、项目的目的和意义": ["简明阐述", "简述"],
        "十二、总结": ["项目经理从技术", "从技术", "总结该项目"],
    }

    def _parse_text_areas(self, doc: ProposalDocument, secs: dict[str, tuple[int, int]]):
        """抽取文本区域模块的内容：取模块内（排除标题/指引行）最长文本单元格。"""
        for module, hints in self._TEXT_MODULES.items():
            rng = secs.get(module)
            if not rng:
                continue
            lo, hi = rng
            best_text, best_coord = "", ""
            for row in self.ws.iter_rows(min_row=lo, max_row=hi):
                for cell in row:
                    t = _norm(cell.value)
                    if not t or t == module or t.startswith(module[:3]):
                        continue
                    if any(t.startswith(h) for h in hints):   # 填写指引，跳过
                        continue
                    if len(t) > len(best_text):
                        best_text, best_coord = t, cell.coordinate
            if best_text:
                doc.text_areas[module] = FieldValue(
                    key=module, value=best_text, status=ExtractStatus.EXTRACTED,
                    anchor=module, source_cell=best_coord)
            else:
                doc.text_areas[module] = FieldValue(
                    key=module, status=ExtractStatus.MISSING, anchor=module,
                    reason="模块内无有效文本内容（仅标题/指引）")

    # ---------- 模块一：人员安排表 ----------
    def _parse_personnel_table(self, doc: ProposalDocument, rng: tuple[int, int] | None):
        """解析人员安排表：分类/人力投入（人月）各行 + 合计行。"""
        if rng is None:
            return
        lo, hi = rng
        header = self.find_label("人力投入", (lo, hi))
        if not header:
            return
        hr = coordinate_to_tuple(header)[0]
        cat_col = coordinate_to_tuple(self.find_label("分类", (lo, hi)) or "E9")[1]
        inv_col = coordinate_to_tuple(header)[1]
        rows: list[dict] = []
        total = None
        for r in range(hr + 1, hi + 1):
            cat = _norm(self.ws.cell(row=r, column=cat_col).value)
            inv = self.ws.cell(row=r, column=inv_col).value
            # 合计行：「合计」可能落在 A~F 任一列（本模板在 B 列）
            row_txt = [_norm(self.ws.cell(row=r, column=c).value) for c in range(1, 7)]
            if "合计" in row_txt:
                total = _num(inv)
                break
            if cat:
                rows.append({"分类": cat, "人月": _num(inv)})
        doc.tables["人员安排"] = rows
        if total is not None:
            doc.tables["人员安排合计"] = [{"合计人月": total}]

    def _is_stop_label(self, value) -> bool:
        """该单元格是否是「另一个字段/章节标题」—— 复选框扫描遇到它即停。"""
        t = _norm(value)
        if not t:
            return False
        stops = set(self.INFO_FIELDS.keys()) | set(self.CHECKBOX_FIELDS) | set(SECTION_TITLES)
        return any(t == s or t.startswith(s) for s in stops)

    def _parse_checkbox_row(self, doc: ProposalDocument, label_text: str,
                            rng: tuple[int, int]):
        """解析同一行的复选框：扫 label 所在行，True/False 后跟选项文本；遇到下一 label 停。"""
        coord = self.find_label(label_text, rng)
        if coord is None:
            doc.warnings.append(f"勾选项「{label_text}」标题未命中")
            return
        r, c = coordinate_to_tuple(coord)
        options: list[tuple[str, bool]] = []
        cells = [self.ws.cell(row=r, column=col) for col in range(c + 1, self.ws.max_column + 1)]
        i = 0
        while i < len(cells):
            v = cells[i].value
            if isinstance(v, bool):
                opt = ""
                if i + 1 < len(cells) and _norm(cells[i + 1].value):
                    opt = _norm(cells[i + 1].value)
                options.append((opt, bool(v)))
                i += 2
            elif self._is_stop_label(v):
                break          # 撞到下一个字段标题（如「适用法规/体系」），停止
            else:
                i += 1
        if options:
            doc.checkboxes[label_text] = options
